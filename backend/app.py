import logging as logger
import traceback

from flask import Flask, request, jsonify, current_app, send_from_directory, Response
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime, timedelta, timezone
import os
import uuid
import logging
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from PIL import Image
import io
from werkzeug.security import generate_password_hash
from functools import wraps

from config import config
from models import UserRole, BoardPermission
from postgres_models import db as postgres_db, PostgreSQLDatabaseManager, TermsVersion
from auth import AuthManager, require_auth, require_board_access, require_board_owner, require_board_admin
from b2_storage import create_b2_service

# Load environment variables
load_dotenv()

# File upload configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_image_format(filename):
    """Get image format from filename"""
    if not filename:
        return None
    
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    # Map extensions to formats
    format_map = {
        'jpg': 'JPEG',
        'jpeg': 'JPEG', 
        'png': 'PNG',
        'gif': 'GIF',
        'webp': 'WEBP'
    }
    
    return format_map.get(ext, 'JPEG')  # Default to JPEG

def create_file_storage_wrapper(bytes_io_obj, filename, content_type):
    """
    Create a FileStorage-like wrapper around BytesIO object for B2 compatibility
    Args:
        bytes_io_obj: BytesIO object with compressed image
        filename: Original filename
        content_type: Original content type
    Returns:
        FileStorage-like object
    """
    class FileStorageWrapper:
        def __init__(self, bytes_io, filename, content_type):
            self.bytes_io = bytes_io
            self.filename = filename
            self.content_type = content_type
            self._position = 0
        
        def seek(self, position):
            self._position = position
            self.bytes_io.seek(position)
        
        def read(self, size=None):
            if size is None:
                return self.bytes_io.read()
            return self.bytes_io.read(size)
        
        def tell(self):
            return self.bytes_io.tell()
    
    return FileStorageWrapper(bytes_io_obj, filename, content_type)

def validate_image_file(file, max_file_size=None):
    """
    Validate image file size and type
    Args:
        file: FileStorage object from Flask
        max_file_size: Maximum file size in bytes (optional)
    Returns:
        tuple: (is_valid, error_message)
    """
    # Check file size (max 10MB before compression by default)
    if max_file_size is None:
        max_file_size = 10 * 1024 * 1024  # 10MB default
    
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to beginning
    
    if file_size > max_file_size:
        return False, f"File too large: {file_size / (1024*1024):.1f}MB. Maximum allowed: {max_file_size / (1024*1024):.1f}MB"
    
    # Check if it's a valid image
    try:
        img = Image.open(file)
        img.verify()  # Verify it's actually an image
        file.seek(0)  # Reset after verification
        return True, None
    except Exception as e:
        return False, f"Invalid image file: {str(e)}"

def compress_image(image_file, max_width=1200, max_height=1200, quality=85):
    """
    Compress image to reduce file size while maintaining quality
    Args:
        image_file: FileStorage object from Flask
        max_width: Maximum width in pixels
        max_height: Maximum height in pixels
        quality: JPEG quality (1-100)
    Returns:
        BytesIO object with compressed image and filename attribute
    """
    try:
        # Store original file position
        original_position = image_file.tell()
        
        # Open image
        img = Image.open(image_file)
        
        # Convert to RGB if necessary (for JPEG)
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        
        # Get original dimensions
        original_width, original_height = img.size
        
        # Calculate new dimensions maintaining aspect ratio
        should_resize = original_width > max_width or original_height > max_height
        
        if should_resize:
            ratio = min(max_width / original_width, max_height / original_height)
            new_width = int(original_width * ratio)
            new_height = int(original_height * ratio)
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            print(f"📸 Resized image from {original_width}x{original_height} to {new_width}x{new_height}")
        else:
            print(f"📸 Image already within size limits: {original_width}x{original_height}")
        
        # Check if compression is worth it
        if not should_resize and quality >= 95:
            print(f"📸 Image already optimal - minimal compression applied")
        
        # Save compressed image to BytesIO
        output = io.BytesIO()
        
        # Add filename attribute to BytesIO object for B2 compatibility
        output.filename = image_file.filename
        
        # Determine format and save with appropriate settings
        image_format = get_image_format(image_file.filename)
        
        if image_format == 'PNG':
            # PNG with optimization
            img.save(output, format='PNG', optimize=True, compress_level=9)
        elif image_format == 'WEBP':
            # WebP with quality setting
            img.save(output, format='WEBP', quality=quality, method=6)  # method 6 = best compression
        elif image_format == 'GIF':
            # GIF with optimization
            img.save(output, format='GIF', optimize=True)
        else:
            # JPEG with quality and optimization
            img.save(output, format='JPEG', quality=quality, optimize=True, progressive=True)
        
        output.seek(0)
        
        # Get file size info
        image_file.seek(0, 2)  # Seek to end to get file size
        original_size = image_file.tell()
        image_file.seek(original_position)  # Reset to original position
        
        compressed_size = len(output.getvalue())
        compression_ratio = (1 - compressed_size / original_size) * 100 if original_size > 0 else 0
        
        print(f"📸 File size: {original_size / 1024:.1f}KB -> {compressed_size / 1024:.1f}KB ({compression_ratio:.1f}% reduction)")
        
        # Log additional compression info
        if compression_ratio > 0:
            print(f"📸 ✅ Compression successful! Saved {compression_ratio:.1f}% of storage space")
        else:
            print(f"📸 ⚠️  No compression achieved - file size increased")
        
        # Log final image info
        final_size_mb = compressed_size / (1024 * 1024)
        if final_size_mb > 1:
            print(f"📸 ⚠️  Final image size: {final_size_mb:.2f}MB - consider reducing quality further")
        else:
            print(f"📸 ✅ Final image size: {final_size_mb:.2f}MB - good compression")
        
        # Log compression summary
        print(f"📸 📊 Compression Summary:")
        print(f"   - Original: {original_width}x{original_height}, {original_size / 1024:.1f}KB")
        print(f"   - Final: {img.size[0]}x{img.size[1]}, {compressed_size / 1024:.1f}KB")
        print(f"   - Savings: {compression_ratio:.1f}%")
        
        # Log format-specific info
        print(f"   - Format: {image_format}")
        print(f"   - Quality: {quality}")
        
        # Log storage savings
        if compression_ratio > 50:
            print(f"   - 🎉 Excellent compression! Saved more than 50%")
        elif compression_ratio > 25:
            print(f"   - ✅ Good compression! Saved more than 25%")
        elif compression_ratio > 0:
            print(f"   - 👍 Some compression achieved")
        else:
            print(f"   - ⚠️  No compression - consider adjusting settings")
        
        # Log recommendations
        if final_size_mb > 2:
            print(f"   - 💡 Tip: Consider reducing quality to 75 for smaller files")
        elif final_size_mb > 1:
            print(f"   - 💡 Tip: Consider reducing quality to 80 for smaller files")
        
        # Log final status
        print(f"📸 🚀 Image compression completed successfully!")
        
        return output
        
    except Exception as e:
        print(f"❌ Error compressing image: {str(e)}")
        # Return original file if compression fails
        image_file.seek(0)
        return image_file

def create_upload_folders():
    """Create upload directories if they don't exist"""
    upload_dir = os.path.join(os.getcwd(), UPLOAD_FOLDER)
    expense_images_dir = os.path.join(upload_dir, 'expense_images')
    
    os.makedirs(upload_dir, exist_ok=True)
    os.makedirs(expense_images_dir, exist_ok=True)
    
    return upload_dir, expense_images_dir

def normalize_expense_date(expense_date):
    """Normalize expense date to ensure timezone-aware datetime object"""
    try:
        # Handle datetime objects
        if isinstance(expense_date, datetime):
            if expense_date.tzinfo is None:
                return expense_date.replace(tzinfo=timezone.utc)
            return expense_date
        
        # Handle string dates
        expense_date_str = str(expense_date)
        
        # If already has timezone info
        if 'Z' in expense_date_str:
            return datetime.fromisoformat(expense_date_str.replace('Z', '+00:00'))
        elif '+' in expense_date_str and (':' in expense_date_str.split('+')[-1]):
            return datetime.fromisoformat(expense_date_str)
        else:
            # If no timezone info, treat as UTC
            dt = datetime.fromisoformat(expense_date_str)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
    except Exception as e:
        print(f"Error normalizing date {expense_date}: {e}")
        # Return current time as fallback
        return datetime.now(timezone.utc)


# Budget helper functions - DISABLED
def check_budget_alerts(board_id, board_budget_amount, board_budget_alerts, total_expenses):
    """
    Check if any budget alerts should be triggered - return only the highest triggered alert
    
    TODO: לוגיקת התקציב מושבתת זמנית - צריך עבודה נוספת
    """
    # DISABLED: Budget alerts temporarily disabled
    return []
    
    # COMMENTED OUT - Budget alerts logic (needs work)
    # if not board_budget_amount or not board_budget_alerts:
    #     return []
    # 
    # expense_percentage = (total_expenses / board_budget_amount) * 100
    # highest_triggered_alert = None
    # 
    # # Find the highest percentage that was triggered
    # for alert_percentage in sorted(board_budget_alerts, reverse=True):  # Sort descending
    #     if expense_percentage >= alert_percentage:
    #         highest_triggered_alert = {
    #             'percentage': alert_percentage,
    #             'current_percentage': expense_percentage,
    #             'budget_amount': board_budget_amount,
    #             'current_expenses': total_expenses
    #         }
    #         break  # Take only the highest triggered alert
    # 
    # return [highest_triggered_alert] if highest_triggered_alert else []


def create_app(config_name='default'):
    """Application factory pattern"""
    app = Flask(__name__)


    
    
    # Load configuration
    app.config.from_object(config[config_name])
    
    # Initialize extensions
    CORS(app, 
         origins=app.config['CORS_ORIGINS'],
         supports_credentials=True,
         allow_headers=['Content-Type', 'Authorization'],
         methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])
    
    print("🔧 Backend: Initializing JWT with secret:", app.config['JWT_SECRET_KEY'][:10] + "...")
    jwt = JWTManager(app)
    
    # Temporarily disable rate limiter to fix the issue
    # limiter = Limiter(
    #     app=app,
    #     key_func=get_remote_address,
    #     default_limits=app.config['RATELIMIT_DEFAULT'],
    #     storage_uri=app.config['RATELIMIT_STORAGE_URL']
    # )
    
    # Initialize PostgreSQL database
    print("🔧 Backend: Using PostgreSQL database")
    
    # Additional fix for Heroku DATABASE_URL (postgres:// -> postgresql://)
    if app.config.get('DATABASE_URL') and app.config.get('DATABASE_URL').startswith('postgres://'):
        fixed_url = app.config.get('DATABASE_URL').replace('postgres://', 'postgresql://', 1)
        app.config['DATABASE_URL'] = fixed_url
        app.config['SQLALCHEMY_DATABASE_URI'] = fixed_url
        print(f"🔧 Fixed DATABASE_URL: {fixed_url}")
    
    # Additional fix for SQLALCHEMY_DATABASE_URI
    if app.config.get('SQLALCHEMY_DATABASE_URI') and app.config.get('SQLALCHEMY_DATABASE_URI').startswith('postgres://'):
        fixed_uri = app.config.get('SQLALCHEMY_DATABASE_URI').replace('postgres://', 'postgresql://', 1)
        app.config['SQLALCHEMY_DATABASE_URI'] = fixed_uri
        print(f"🔧 Fixed SQLALCHEMY_DATABASE_URI: {fixed_uri}")
    
    # Additional fix for SQLAlchemy engine options
    if not app.config.get('SQLALCHEMY_ENGINE_OPTIONS'):
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {}
    
    if 'connect_args' not in app.config['SQLALCHEMY_ENGINE_OPTIONS']:
        app.config['SQLALCHEMY_ENGINE_OPTIONS']['connect_args'] = {}
    
    if 'sslmode' not in app.config['SQLALCHEMY_ENGINE_OPTIONS']['connect_args']:
        app.config['SQLALCHEMY_ENGINE_OPTIONS']['connect_args']['sslmode'] = 'require'
    
    postgres_db.init_app(app)
    
    # Create tables and initialize data
    with app.app_context():
        postgres_db.create_all()
        db_manager = PostgreSQLDatabaseManager(app)
        db_manager.initialize_default_data()
        db_manager.cleanup_expired_pending_registrations()
        # Migrate existing debts to new format
        db_manager.migrate_existing_debts_to_new_format()
    
    app.db_manager = db_manager
    
    # Budget helper functions (moved inside create_app to access db_manager)
    def format_datetime_field(dt_field):
        """Helper function to safely format datetime fields that might be strings or datetime objects"""
        if not dt_field:
            return None
        if hasattr(dt_field, 'isoformat'):
            return dt_field.isoformat()
        return dt_field  # Already a string
    
    def format_time_field(time_field):
        """Helper function to safely format time fields for JSON serialization"""
        if not time_field:
            return None
        if hasattr(time_field, 'strftime'):
            return time_field.strftime('%H:%M')
        return str(time_field)  # Already a string
    
    def check_and_reset_budget_if_needed(board_id):
        """
        בדוק אם צריך לאפס את התקציב של הלוח ואפס אותו אם צריך.
        מחזיר True אם התקציב אופס, False אחרת.
        
        TODO: לוגיקת התקציב מושבתת זמנית - צריך עבודה נוספת
        """
        # DISABLED: Budget logic temporarily disabled
        return False
        
        # COMMENTED OUT - Budget reset logic (needs work)
        # try:
        #     board = db_manager.get_board_by_id(board_id)
        #     if not board or not board.budget_auto_reset or not board.budget_reset_day:
        #         return False
        #     
        #     from datetime import datetime, timezone, time
        #     
        #     now = datetime.now(timezone.utc)
        #     reset_day = board.budget_reset_day
        #     
        #     # בנה את תאריך ושעת האיפוס הבא
        #     reset_datetime = _get_next_reset_datetime(now, reset_day, board.budget_reset_time)
        #     
        #     # בדוק אם יש איפוס אחרון
        #     if board.budget_last_reset is None:
        #         # מעולם לא אופס - אפס אם עבר זמן האיפוס הבא
        #         should_reset = now >= reset_datetime
        #         print(f"💰 First time reset check: now={now.strftime('%Y-%m-%d %H:%M')}, reset_time={reset_datetime.strftime('%Y-%m-%d %H:%M')}, should_reset={should_reset}")
        #     else:
        #         # המר את תאריך האיפוס האחרון
        #         if isinstance(board.budget_last_reset, str):
        #             last_reset = datetime.fromisoformat(board.budget_last_reset.replace('Z', '+00:00'))
        #         else:
        #             last_reset = board.budget_last_reset.replace(tzinfo=timezone.utc) if board.budget_last_reset.tzinfo is None else board.budget_last_reset
        #         
        #         # אפס רק אם עבר זמן האיפוס מאז האיפוס האחרון
        #         should_reset = now >= reset_datetime and last_reset < reset_datetime
        #         print(f"💰 Reset check: now={now.strftime('%Y-%m-%d %H:%M')}, reset_time={reset_datetime.strftime('%Y-%m-%d %H:%M')}, last_reset={last_reset.strftime('%Y-%m-%d %H:%M')}, should_reset={should_reset}")
        #     
        #     if should_reset:
        #         reset_time_str = f" at {board.budget_reset_time}" if board.budget_reset_time else ""
        #         print(f"💰 Resetting budget for board {board_id} - auto reset day: {reset_day}{reset_time_str}")
        #         
        #         # עדכן את תאריך האיפוס האחרון לזמן הנוכחי
        #         update_data = {
        #             'budget_last_reset': now
        #         }
        #         
        #         success = db_manager.update_board(board_id, update_data)
        #         if success:
        #             print(f"💰 Budget reset successful for board {board_id}")
        #             return True
        #         else:
        #             print(f"💰 Failed to update budget reset date for board {board_id}")
        #             return False
        #     
        #     return False
        #     
        # except Exception as e:
        #     print(f"💰 Error checking/resetting budget for board {board_id}: {e}")
        #     import traceback
        #     traceback.print_exc()
        #     return False
    
    def _get_next_reset_datetime(current_time, reset_day, reset_time_str):
        """
        חשב את תאריך ושעת האיפוס הבא בהתבסס על היום והשעה שנקבעו
        
        TODO: לוגיקת התקציב מושבתת זמנית - צריך עבודה נוספת
        """
        # DISABLED: Budget logic temporarily disabled
        from datetime import datetime, timezone
        return datetime.now(timezone.utc)
        
        # COMMENTED OUT - Budget reset datetime calculation (needs work)
        # from datetime import datetime, timezone, time
        # import calendar
        # 
        # # קבע את השעה (ברירת מחדל 09:00)
        # reset_hour, reset_minute = 9, 0
        # if reset_time_str:
        #     try:
        #         reset_hour, reset_minute = map(int, reset_time_str.split(':'))
        #     except (ValueError, AttributeError):
        #         pass  # השתמש בברירת המחדל
        # 
        # reset_time = time(reset_hour, reset_minute)
        # 
        # # התחל מהחודש הנוכחי
        # year = current_time.year
        # month = current_time.month
        # 
        # # וודא שיום האיפוס תקין לחודש הנוכחי
        # max_day = calendar.monthrange(year, month)[1]
        # actual_reset_day = min(reset_day, max_day)
        # 
        # # בנה את תאריך האיפוס בחודש הנוכחי
        # reset_datetime = datetime(year, month, actual_reset_day, reset_hour, reset_minute, tzinfo=timezone.utc)
        # 
        # # אם תאריך האיפוס כבר עבר בחודש הנוכחי, עבור לחודש הבא
        # if reset_datetime <= current_time:
        #     # עבור לחודש הבא
        #     if month == 12:
        #         year += 1
        #         month = 1
        #     else:
        #         month += 1
        #     
        #     # וודא שיום האיפוס תקין לחודש הבא
        #     max_day = calendar.monthrange(year, month)[1]
        #     actual_reset_day = min(reset_day, max_day)
        #     
        #     reset_datetime = datetime(year, month, actual_reset_day, reset_hour, reset_minute, tzinfo=timezone.utc)
        # 
        # return reset_datetime

    def calculate_board_total_expenses(board_id):
        """
        Calculate total expenses for a board since last budget reset (if any)
        
        TODO: לוגיקת התקציב מושבתת זמנית - צריך עבודה נוספת
        """
        try:
            board = db_manager.get_board_by_id(board_id)
            if not board:
                return 0.0
            
            # DISABLED: Budget reset logic temporarily disabled
            # Just return total of all expenses for now
            expenses = db_manager.get_board_expenses(board_id)
            total_amount = sum(expense.amount for expense in expenses)
            print(f"💰 Board {board_id}: Total expenses (budget logic disabled): {total_amount} ({len(expenses)} expenses)")
            return total_amount
            
            # COMMENTED OUT - Budget-aware expense calculation (needs work)
            # # בדוק אם צריך איפוס לפני חישוב ההוצאות
            # was_reset = check_and_reset_budget_if_needed(board_id)
            # if was_reset:
            #     print(f"💰 Budget was just reset for board {board_id}, returning 0 expenses")
            #     return 0.0
            # 
            # expenses = db_manager.get_board_expenses(board_id)
            # 
            # # אם יש איפוס אוטומטי ותאריך איפוס אחרון, חשב רק הוצאות מאז האיפוס
            # if board.budget_auto_reset and board.budget_last_reset:
            #     from datetime import datetime, timezone
            #     
            #     # המר את תאריך האיפוס האחרון
            #     if isinstance(board.budget_last_reset, str):
            #         last_reset = datetime.fromisoformat(board.budget_last_reset.replace('Z', '+00:00'))
            #     else:
            #         last_reset = board.budget_last_reset.replace(tzinfo=timezone.utc) if board.budget_last_reset.tzinfo is None else board.budget_last_reset
            #     
            #     # סנן הוצאות רק מאז האיפוס האחרון
            #     filtered_expenses = []
            #     for expense in expenses:
            #         expense_date = normalize_expense_date(expense.date)
            #         if expense_date >= last_reset:
            #             filtered_expenses.append(expense)
            #     
            #     total_amount = sum(expense.amount for expense in filtered_expenses)
            #     print(f"💰 Board {board_id}: Total expenses since last reset ({last_reset.strftime('%Y-%m-%d %H:%M')}): {total_amount} ({len(filtered_expenses)} out of {len(expenses)} expenses)")
            #     return total_amount
            # else:
            #     # אין איפוס אוטומטי - חשב את כל ההוצאות
            #     total_amount = sum(expense.amount for expense in expenses)
            #     print(f"💰 Board {board_id}: Total expenses (no auto-reset): {total_amount} ({len(expenses)} expenses)")
            #     return total_amount
        except Exception as e:
            print(f"Error calculating board total expenses: {e}")
            return 0.0

    def get_budget_status(board_id):
        """
        Get budget status for a board including alerts
        
        TODO: לוגיקת התקציב מושבתת זמנית - צריך עבודה נוספת
        """
        try:
            print(f"💰 Getting budget status for board: {board_id} (BUDGET LOGIC DISABLED)")
            
            board = db_manager.get_board_by_id(board_id)
            
            if not board or not board.budget_amount:
                print(f"💰 No budget found for board {board_id}")
                return {
                    'has_budget': False,
                    'budget_amount': None,
                    'current_expenses': 0.0,
                    'percentage_used': 0.0,
                    'alerts': [],
                    'triggered_alerts': []
                }
            
            # DISABLED: Just show basic budget info without alerts or reset logic
            total_expenses = calculate_board_total_expenses(board_id)
            percentage_used = (total_expenses / board.budget_amount) * 100 if board.budget_amount > 0 else 0
            
            result = {
                'has_budget': True,
                'budget_amount': board.budget_amount,
                'current_expenses': total_expenses,
                'percentage_used': percentage_used,
                'alerts': board.budget_alerts or [],
                'triggered_alerts': [],  # DISABLED: No alerts for now
                'was_reset': False  # DISABLED: No reset logic for now
            }
            print(f"💰 Budget status (alerts disabled): {result}")
            return result
            
            # COMMENTED OUT - Full budget status with alerts and reset logic (needs work)
            # # בדוק ואפס תקציב אם צריך בהתחלה
            # was_reset = check_and_reset_budget_if_needed(board_id)
            # if was_reset:
            #     print(f"💰 Budget was just reset for board {board_id}")
            # 
            # board = db_manager.get_board_by_id(board_id)
            # print(f"💰 Board data: {board}")
            # 
            # if board:
            #     print(f"💰 Board budget_amount: {board.budget_amount}")
            #     print(f"💰 Board budget_alerts: {board.budget_alerts}")
            # 
            # if not board or not board.budget_amount:
            #     print(f"💰 No budget found for board {board_id}")
            #     return {
            #         'has_budget': False,
            #         'budget_amount': None,
            #         'current_expenses': 0.0,
            #         'percentage_used': 0.0,
            #         'alerts': [],
            #         'triggered_alerts': []
            #     }
            # 
            # # חשב הוצאות (כבר לוקח בחשבון איפוס)
            # total_expenses = calculate_board_total_expenses(board_id)
            # print(f"💰 Total expenses: {total_expenses}")
            # percentage_used = (total_expenses / board.budget_amount) * 100 if board.budget_amount > 0 else 0
            # print(f"💰 Percentage used: {percentage_used}%")
            # 
            # # אם התקציב אופס זה עתה, לא צריך התראות
            # if was_reset:
            #     print(f"💰 Budget was reset - no alerts triggered")
            #     triggered_alerts = []
            # else:
            #     triggered_alerts = check_budget_alerts(
            #         board_id, 
            #         board.budget_amount, 
            #         board.budget_alerts or [], 
            #         total_expenses
            #     )
            #     
            #     # Add budget exceeded alert if over 100%
            #     if percentage_used >= 100:
            #         print(f"💰 Budget exceeded! Replacing all alerts with exceeded alert...")
            #         # Clear all previous alerts and show only the exceeded alert
            #         triggered_alerts = [{
            #             'percentage': 100,
            #             'current_percentage': percentage_used,
            #             'budget_amount': board.budget_amount,
            #             'current_expenses': total_expenses,
            #             'is_exceeded': True  # Special flag for exceeded budget
            #         }]
            #         print(f"💰 Showing only exceeded alert")
            # 
            # print(f"💰 Triggered alerts: {triggered_alerts}")
            # 
            # result = {
            #     'has_budget': True,
            #     'budget_amount': board.budget_amount,
            #     'current_expenses': total_expenses,
            #     'percentage_used': percentage_used,
            #     'alerts': board.budget_alerts or [],
            #     'triggered_alerts': triggered_alerts,
            #     'was_reset': was_reset  # הוסף מידע על איפוס
            # }
            # print(f"💰 Final budget status: {result}")
            # return result
        except Exception as e:
            print(f"Error getting budget status for board {board_id}: {e}")
            import traceback
            traceback.print_exc()
            return {
                'has_budget': False,
                'budget_amount': None,
                'current_expenses': 0.0,
                'percentage_used': 0.0,
                'alerts': [],
                'triggered_alerts': []
            }
    
    # Initialize auth
    auth_manager = AuthManager(db_manager)
    app.auth_manager = auth_manager
    
    # Initialize storage services
    upload_dir, expense_images_dir = create_upload_folders()
    app.config['UPLOAD_FOLDER'] = upload_dir
    app.config['EXPENSE_IMAGES_FOLDER'] = expense_images_dir
    
    # Initialize B2 storage service if configured
    b2_service = None
    if app.config.get('UPLOAD_METHOD') == 'b2':
        b2_service = create_b2_service(app.config)
        if b2_service:
            print("🚀 B2 Storage Service initialized successfully")
        else:
            print("⚠️  B2 Storage Service failed to initialize, falling back to local storage")
    
    app.b2_service = b2_service
    
    # File upload function
    def upload_expense_image(file, user_id):
        """Upload expense image and return the file URL"""
        if not file or not allowed_file(file.filename):
            return None
        
        print(f"📸 Processing image upload: {file.filename}")
        
        # Validate image file
        is_valid, error_message = validate_image_file(
            file, 
            max_file_size=app.config.get('IMAGE_MAX_FILE_SIZE', 10 * 1024 * 1024)
        )
        if not is_valid:
            print(f"❌ Image validation failed: {error_message}")
            return None
        
        # Compress image before upload
        compressed_file = compress_image(
            file, 
            max_width=app.config.get('IMAGE_MAX_WIDTH', 1200),
            max_height=app.config.get('IMAGE_MAX_HEIGHT', 1200),
            quality=app.config.get('IMAGE_QUALITY', 85)
        )
        
        # Try B2 storage first if available and configured
        if app.config.get('UPLOAD_METHOD') == 'b2' and app.b2_service:
            try:
                # Create a compressed file wrapper that mimics FileStorage for B2
                compressed_file_wrapper = create_file_storage_wrapper(compressed_file, file.filename, file.content_type)
                
                success, file_url, error = app.b2_service.upload_file(compressed_file_wrapper, 'expense_images', user_id)
                if success:
                    print(f"✅ Successfully uploaded compressed image to B2: {file_url}")
                    return file_url
                else:
                    print(f"❌ B2 upload failed: {error}, falling back to local storage")
            except Exception as e:
                print(f"❌ B2 upload exception: {e}, falling back to local storage")
        
        # Fallback to local storage - use compressed version
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{user_id}_{uuid.uuid4().hex}.{file_extension}"
        
        # Save compressed file locally
        file_path = os.path.join(app.config['EXPENSE_IMAGES_FOLDER'], unique_filename)
        
        # Reset file pointer and save
        compressed_file.seek(0)
        with open(file_path, 'wb') as f:
            f.write(compressed_file.getvalue())
        
        print(f"✅ Saved compressed image locally: {file_path}")
        
        # Return URL path for accessing the image
        return f"/api/uploads/expense_images/{unique_filename}"
    
    # Error handlers
    @app.errorhandler(404)
    def not_found(error):
        return jsonify({'error': 'Resource not found'}), 404
    
    @app.errorhandler(500)
    def internal_error(error):
        return jsonify({'error': 'Internal server error'}), 500
    
    @app.errorhandler(429)
    def ratelimit_handler(error):
        return jsonify({'error': 'Rate limit exceeded'}), 429

    # Health check
    @app.route('/api/health', methods=['GET'])
    def health_check():
        return jsonify({
            'status': 'healthy',
            'message': 'Expense Manager API is running',
            'timestamp': datetime.now().isoformat()
        })

    # Authentication routes
    @app.route('/api/auth/register', methods=['POST'])
    def register():
        print("🔍 === REGULAR REGISTRATION ENDPOINT CALLED ===")
        try:
            data = request.get_json()
            
            # בדיקות תקינות
            if not all(key in data for key in ['email', 'password', 'first_name', 'last_name']):
                return jsonify({'error': 'Missing required fields'}), 400
            
            # בדיקה שהמשתמש מסכים לתנאים
            if not data.get('accepted_terms', False):
                return jsonify({'error': 'You must accept the terms and conditions to register'}), 400
            
            # בדיקה שיש תאריך אישור תנאים
            if not data.get('terms_accepted_at'):
                return jsonify({'error': 'Terms acceptance timestamp is required'}), 400
            
            # בדיקה שהמשתמש לא קיים כבר
            if db_manager.get_user_by_email(data['email']):
                return jsonify({'error': 'Email already registered'}), 400
            
            # Generate username from email
            username = data['email'].split('@')[0]
            
            # יצירת משתמש חדש עם תנאי השימוש
            user_data = {
                'email': data['email'].lower(),
                'username': username,
                'password_hash': generate_password_hash(data['password']),
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'accepted_terms': data.get('accepted_terms', True),
                'terms_accepted_at': data.get('terms_accepted_at', datetime.utcnow()),
                'terms_version_signed': None  # Will be set when user actually accepts terms
            }
            
            # יצירת המשתמש בבסיס הנתונים
            user = db_manager.create_user(user_data)
            
            if user:
                # Check for pending invitations for this email
                print(f"🔍 Checking for pending invitations for email: {data['email'].lower()}")
                pending_invitations = db_manager.get_pending_invitations_by_email(data['email'].lower())
                boards_joined = []
                
                print(f"🔍 Found {len(pending_invitations)} pending invitation(s) for {data['email']}")
                
                if pending_invitations:
                    print(f"🎉 Processing {len(pending_invitations)} pending invitation(s) for {data['email']}")
                    
                    for invitation in pending_invitations:
                        print(f"🔍 Processing invitation {invitation.id} for board {invitation.board_id}")
                        try:
                            # Accept the invitation automatically
                            print(f"🔍 Attempting to accept invitation {invitation.id} for user {user.id}")
                            success = db_manager.accept_invitation(invitation.id, user.id)
                            print(f"🔍 Accept invitation result: {success}")
                            
                            if success:
                                board = db_manager.get_board_by_id(invitation.board_id)
                                if board:
                                    boards_joined.append({
                                        'board_id': board.id,
                                        'board_name': board.name,
                                        'role': invitation.role
                                    })
                                    print(f"✅ Successfully joined user {user.id} to board '{board.name}' with role '{invitation.role}'")
                                    
                                    # Verify membership was created
                                    member_role = db_manager.get_user_role_in_board(user.id, board.id)
                                    print(f"🔍 Verification - User {user.id} role in board {board.id}: {member_role}")
                                else:
                                    print(f"❌ Board {invitation.board_id} not found")
                            else:
                                print(f"❌ Failed to accept invitation {invitation.id}")
                        except Exception as e:
                            print(f"❌ Error accepting invitation {invitation.id}: {e}")
                            import traceback
                            traceback.print_exc()
                else:
                    print(f"ℹ️ No pending invitations found for {data['email']}")
                
                response_data = {
                    'message': 'User registered successfully',
                    'user_id': user.id,
                    'accepted_terms': user.accepted_terms,
                    'terms_accepted_at': user.terms_accepted_at,
                    'terms_version_signed': user.terms_version_signed
                }
                
                # Add information about automatically joined boards
                if boards_joined:
                    response_data['boards_joined'] = boards_joined
                    response_data['message'] += f' and automatically joined {len(boards_joined)} board(s)'
                
                return jsonify(response_data), 201
            else:
                return jsonify({'error': 'Failed to create user'}), 500
            
        except Exception as e:
            print(f"❌ Registration error: {e}")
            return jsonify({'error': 'Registration failed'}), 500

    @app.route('/api/auth/register-with-invitation', methods=['POST'])
    def register_with_invitation():
        """Register user with invitation token"""
        try:
            data = request.get_json()
            
            # בדיקות תקינות
            if not all(key in data for key in ['email', 'password', 'first_name', 'last_name', 'invitation_token']):
                return jsonify({'error': 'Missing required fields'}), 400
            
            # בדיקה שהמשתמש מסכים לתנאים
            if not data.get('accepted_terms', False):
                return jsonify({'error': 'You must accept the terms and conditions to register'}), 400
            
            # בדיקה שיש תאריך אישור תנאים
            if not data.get('terms_accepted_at'):
                return jsonify({'error': 'Terms acceptance timestamp is required'}), 400
            
            # Check if invitation exists and is valid
            invitation = db_manager.get_invitation_by_token(data['invitation_token'])
            if not invitation:
                return jsonify({'error': 'Invalid or expired invitation'}), 400
            
            # Check if invitation is expired
            invitation_expires = datetime.fromisoformat(invitation.expires_at.replace('Z', '+00:00'))
            if invitation_expires < datetime.now(timezone.utc):
                return jsonify({'error': 'Invitation has expired'}), 400
            
            # Check if invitation email matches registration email
            if invitation.email.lower() != data['email'].lower():
                return jsonify({'error': 'Email does not match invitation'}), 400
            
            # Check if invitation is already accepted
            if invitation.accepted_at:
                return jsonify({'error': 'Invitation has already been accepted'}), 400
            
            # בדיקה שהמשתמש לא קיים כבר
            if db_manager.get_user_by_email(data['email']):
                return jsonify({'error': 'Email already registered'}), 400
            
            # Generate username from email
            username = data['email'].split('@')[0]
            
            # יצירת משתמש חדש עם תנאי השימוש
            user_data = {
                'email': data['email'].lower(),
                'username': username,
                'password_hash': generate_password_hash(data['password']),
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'accepted_terms': data.get('accepted_terms', True),
                'terms_accepted_at': data.get('terms_accepted_at', datetime.utcnow()),
                'terms_version_signed': None  # Will be set when user actually accepts terms
            }
            
            # יצירת המשתמש בבסיס הנתונים
            user = db_manager.create_user(user_data)
            
            if user:
                # Accept the specific invitation
                success = db_manager.accept_invitation(invitation.id, user.id)
                board = db_manager.get_board_by_id(invitation.board_id)
                
                if success and board:
                    print(f"✅ User {user.id} registered and joined board '{board.name}' via invitation")
                    
                    # Check for any other pending invitations for this email
                    other_invitations = db_manager.get_pending_invitations_by_email(data['email'].lower())
                    boards_joined = [{
                        'board_id': board.id,
                        'board_name': board.name,
                        'role': invitation.role
                    }]
                    
                    # Accept other pending invitations automatically
                    for other_invitation in other_invitations:
                        if other_invitation.id != invitation.id:  # Skip the one we already processed
                            try:
                                success = db_manager.accept_invitation(other_invitation.id, user.id)
                                if success:
                                    other_board = db_manager.get_board_by_id(other_invitation.board_id)
                                    if other_board:
                                        boards_joined.append({
                                            'board_id': other_board.id,
                                            'board_name': other_board.name,
                                            'role': other_invitation.role
                                        })
                                        print(f"✅ Also joined user {user.id} to board '{other_board.name}'")
                            except Exception as e:
                                print(f"❌ Error accepting additional invitation {other_invitation.id}: {e}")
                    
                    return jsonify({
                        'message': f'User registered successfully and joined {len(boards_joined)} board(s)',
                        'user_id': user.id,
                        'accepted_terms': user.accepted_terms,
                        'terms_accepted_at': user.terms_accepted_at,
                        'terms_version_signed': user.terms_version_signed,
                        'boards_joined': boards_joined
                    }), 201
                else:
                    return jsonify({'error': 'Failed to accept invitation'}), 500
            else:
                return jsonify({'error': 'Failed to create user'}), 500
            
        except Exception as e:
            print(f"❌ Registration with invitation error: {e}")
            return jsonify({'error': 'Registration failed'}), 500

    @app.route('/api/auth/invitation/<invitation_token>', methods=['GET'])
    def get_invitation_info(invitation_token):
        """Get invitation information by token"""
        try:
            invitation = db_manager.get_invitation_by_token(invitation_token)
            if not invitation:
                return jsonify({'error': 'Invalid or expired invitation'}), 404
            
            # Check if invitation is expired
            invitation_expires = datetime.fromisoformat(invitation.expires_at.replace('Z', '+00:00'))
            if invitation_expires < datetime.now(timezone.utc):
                return jsonify({'error': 'Invitation has expired'}), 400
            
            # Check if invitation is already accepted
            if invitation.accepted_at:
                return jsonify({'error': 'Invitation has already been accepted'}), 400
            
            # Get board information
            board = db_manager.get_board_by_id(invitation.board_id)
            if not board:
                return jsonify({'error': 'Board not found'}), 404
            
            # Get inviter information
            inviter = db_manager.get_user_by_id(invitation.invited_by)
            inviter_name = f"{inviter.first_name} {inviter.last_name}" if inviter else "משתמש לא ידוע"
            
            return jsonify({
                'valid': True,
                'invitation': {
                    'id': invitation.id,
                    'email': invitation.email,
                    'role': invitation.role,
                    'expires_at': invitation.expires_at,
                    'created_at': invitation.created_at
                },
                'board': {
                    'id': board.id,
                    'name': board.name,
                    'description': board.description,
                    'currency': board.currency
                },
                'inviter': {
                    'name': inviter_name
                }
            }), 200
            
        except Exception as e:
            print(f"❌ Error getting invitation info: {e}")
            return jsonify({'error': 'Failed to get invitation information'}), 500

    @app.route('/download', methods=['GET'])
    def download_page():
        """App download page with invitation tracking"""
        invitation_token = request.args.get('invitation')
        
        # Get invitation info if token provided
        invitation_info = None
        if invitation_token:
            try:
                invitation = db_manager.get_invitation_by_token(invitation_token)
                if invitation:
                    # Check if invitation is valid
                    invitation_expires = datetime.fromisoformat(invitation.expires_at.replace('Z', '+00:00'))
                    if invitation_expires >= datetime.now(timezone.utc) and not invitation.accepted_at:
                        board = db_manager.get_board_by_id(invitation.board_id)
                        inviter = db_manager.get_user_by_id(invitation.invited_by)
                        inviter_name = f"{inviter.first_name} {inviter.last_name}" if inviter else "משתמש"
                        
                        invitation_info = {
                            'email': invitation.email,
                            'board_name': board.name if board else 'לוח לא ידוע',
                            'inviter_name': inviter_name
                        }
            except Exception as e:
                print(f"Error getting invitation info for download page: {e}")
        
        # Create HTML download page
        html_content = f"""
        <!DOCTYPE html>
        <html dir="rtl" lang="he">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Homis - הורדת האפליקציה</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    margin: 0;
                    padding: 20px;
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }}
                .container {{
                    background: white;
                    border-radius: 20px;
                    padding: 40px;
                    max-width: 500px;
                    width: 100%;
                    box-shadow: 0 20px 40px rgba(0,0,0,0.1);
                    text-align: center;
                }}
                .logo {{
                    font-size: 48px;
                    color: #3498db;
                    margin-bottom: 10px;
                }}
                .app-name {{
                    font-size: 32px;
                    font-weight: bold;
                    color: #2c3e50;
                    margin-bottom: 10px;
                }}
                .tagline {{
                    color: #7f8c8d;
                    font-size: 16px;
                    margin-bottom: 30px;
                }}
                .invitation-box {{
                    background: #e8f4fd;
                    border-radius: 12px;
                    padding: 20px;
                    margin-bottom: 30px;
                    border-right: 4px solid #3498db;
                }}
                .invitation-title {{
                    font-size: 18px;
                    font-weight: bold;
                    color: #2c3e50;
                    margin-bottom: 10px;
                }}
                .invitation-details {{
                    color: #555;
                    line-height: 1.6;
                }}
                .steps {{
                    background: #f8f9fa;
                    border-radius: 12px;
                    padding: 25px;
                    margin-bottom: 30px;
                    text-align: right;
                }}
                .steps h3 {{
                    color: #2c3e50;
                    margin-bottom: 20px;
                    text-align: center;
                }}
                .step {{
                    display: flex;
                    align-items: center;
                    margin-bottom: 15px;
                    padding: 10px;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .step-number {{
                    background: #3498db;
                    color: white;
                    width: 30px;
                    height: 30px;
                    border-radius: 50%;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    font-weight: bold;
                    margin-left: 15px;
                    flex-shrink: 0;
                }}
                .step-text {{
                    color: #555;
                    line-height: 1.5;
                }}
                .download-buttons {{
                    display: flex;
                    gap: 15px;
                    justify-content: center;
                    margin-bottom: 20px;
                    flex-wrap: wrap;
                }}
                .download-btn {{
                    display: inline-flex;
                    align-items: center;
                    background: #2c3e50;
                    color: white;
                    text-decoration: none;
                    padding: 15px 25px;
                    border-radius: 12px;
                    font-weight: bold;
                    transition: transform 0.2s;
                    min-width: 200px;
                    justify-content: center;
                }}
                .download-btn:hover {{
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(0,0,0,0.2);
                }}
                .download-btn.ios {{
                    background: #007AFF;
                }}
                .download-btn.android {{
                    background: #34A853;
                }}
                .store-icon {{
                    width: 24px;
                    height: 24px;
                    margin-left: 10px;
                }}
                .footer {{
                    color: #7f8c8d;
                    font-size: 14px;
                    margin-top: 30px;
                }}
                @media (max-width: 600px) {{
                    .container {{ padding: 20px; }}
                    .download-buttons {{ flex-direction: column; }}
                    .download-btn {{ min-width: auto; }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">💰</div>
                <div class="app-name">Homis</div>
                <div class="tagline">ניהול הוצאות משותפות</div>
                
                {f'''
                <div class="invitation-box">
                    <div class="invitation-title">🎉 הוזמנת להצטרף!</div>
                    <div class="invitation-details">
                        <strong>{invitation_info['inviter_name']}</strong> הזמין אותך להצטרף ללוח הוצאות "<strong>{invitation_info['board_name']}</strong>"<br>
                        עם המייל: <strong style="color: #3498db;">{invitation_info['email']}</strong>
                    </div>
                </div>
                ''' if invitation_info else ''}
                
                <div class="steps">
                    <h3>💫 איך להתחיל?</h3>
                    <div class="step">
                        <div class="step-number">1</div>
                        <div class="step-text">הורד את אפליקציית Homis מחנות האפליקציות</div>
                    </div>
                    <div class="step">
                        <div class="step-number">2</div>
                        <div class="step-text">
                            פתח את האפליקציה וצור משתמש חדש
                            {f'עם המייל: <strong style="color: #3498db;">{invitation_info["email"]}</strong>' if invitation_info else ''}
                        </div>
                    </div>
                    <div class="step">
                        <div class="step-number">3</div>
                        <div class="step-text">
                            {f'הלוח "{invitation_info["board_name"]}" יופיע לך אוטומטית!' if invitation_info else 'התחל לנהל את ההוצאות שלך!'}
                        </div>
                    </div>
                </div>
                
                <div class="download-buttons">
                    <a href="https://apps.apple.com/app/homis" class="download-btn ios">
                        📱 App Store
                    </a>
                    <a href="https://play.google.com/store/apps/details?id=com.homis" class="download-btn android">
                        📱 Google Play
                    </a>
                </div>
                
                <div class="footer">
                    אפליקציית Homis מאפשרת לך ולחבריך לעקוב אחר הוצאות משותפות, 
                    לחלוק עלויות בצורה הוגנת ולנהל חובות בקלות.
                </div>
            </div>
        </body>
        </html>
        """
        
        return Response(html_content, mimetype='text/html')

    @app.route('/api/auth/login', methods=['POST'])
    def login():
        """Login user"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        result = auth_manager.login_user(data)
        
        if result['valid']:
            return jsonify(result), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    @app.route('/api/auth/refresh', methods=['POST'])
    @jwt_required(refresh=True)
    def refresh():
        """Refresh access token"""
        result = auth_manager.refresh_token()
        
        if result['valid']:
            return jsonify(result), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    @app.route('/api/auth/me', methods=['GET'])
    @require_auth
    def get_current_user():
        """Get current user information"""
        result = auth_manager.get_current_user()
        
        if result['valid']:
            return jsonify(result), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    @app.route('/api/auth/change-password', methods=['POST'])
    @require_auth
    def change_password():
        """Change user password"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_user_id = auth_manager.get_current_user()['user']['id']
        result = auth_manager.change_password(
            current_user_id,
            data.get('current_password'),
            data.get('new_password')
        )
        
        if result['valid']:
            return jsonify({'message': 'Password changed successfully'}), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    # Email verification endpoints
    @app.route('/api/auth/send-verification', methods=['POST'])
    def send_verification():
        """Send verification code for registration"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        result = auth_manager.send_verification_code(data)
        
        if result['valid']:
            return jsonify({'message': result['message']}), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    @app.route('/api/auth/verify-and-register', methods=['POST'])
    def verify_and_register():
        """Verify code and complete registration"""
        print("🔍 === EMAIL VERIFICATION REGISTRATION ENDPOINT CALLED ===")
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        email = data.get('email')
        code = data.get('code')
        
        if not email or not code:
            return jsonify({'error': 'Email and verification code are required'}), 400
        
        # Convert email to lowercase
        email = email.lower()
        
        result = auth_manager.verify_code_and_register(email, code)
        
        if result['valid']:
            return jsonify(result), 201
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    # Password reset endpoints
    @app.route('/api/auth/request-password-reset', methods=['POST'])
    def request_password_reset():
        """Request password reset by sending verification code"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        email = data.get('email')
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        # Convert email to lowercase
        email = email.lower()
        
        result = auth_manager.request_password_reset(email)
        
        if result['valid']:
            return jsonify({'message': result['message']}), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    @app.route('/api/auth/verify-reset-code', methods=['POST'])
    def verify_reset_code():
        """Verify password reset code"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        email = data.get('email')
        code = data.get('code')
        
        if not email or not code:
            return jsonify({'error': 'Email and verification code are required'}), 400
        
        # Convert email to lowercase
        email = email.lower()
        
        result = auth_manager.verify_reset_code(email, code)
        
        if result['valid']:
            return jsonify({'message': result['message']}), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    @app.route('/api/auth/reset-password', methods=['POST'])
    def reset_password():
        """Reset password with verified code"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        email = data.get('email')
        code = data.get('code')
        new_password = data.get('new_password')
        
        if not email or not code or not new_password:
            return jsonify({'error': 'Email, code, and new password are required'}), 400
        
        # Convert email to lowercase
        email = email.lower()
        
        result = auth_manager.reset_password(email, code, new_password)
        
        if result['valid']:
            return jsonify({'message': result['message']}), 200
        else:
            return jsonify({'error': result['error']}), result.get('code', 400)

    # Board routes
    @app.route('/api/boards', methods=['GET', 'OPTIONS'])
    @require_auth
    def get_user_boards():
        """Get all boards for the current user"""
        try:
            current_user_id = auth_manager.get_current_user()['user']['id']
            boards = db_manager.get_user_boards(current_user_id)
            
            board_list = []
            for board in boards:
                # לא צריך לבדוק איפוס כאן - זה יקרה בget_budget_status או calculate_board_total_expenses
                board_dict = {
                    'id': board.id,
                    'name': board.name,
                    'description': board.description,
                    'owner_id': board.owner_id,
                    'created_at': board.created_at,
                    'updated_at': board.updated_at,
                    'currency': board.currency,
                    'timezone': board.timezone,
                    'board_type': board.board_type,
                    'budget_amount': board.budget_amount,
                    'budget_alerts': board.budget_alerts or [],
                    'budget_auto_reset': board.budget_auto_reset,
                    'budget_reset_day': board.budget_reset_day,
                    'budget_reset_time': format_time_field(board.budget_reset_time),
                    'budget_last_reset': format_datetime_field(board.budget_last_reset),
                    'is_default_board': getattr(board, 'is_default_board', False),
                    'user_role': getattr(board, 'user_role', 'member')
                }
                
                # Get member count for each board
                members = db_manager.get_board_members(board.id)
                board_dict['member_count'] = len(members)
                
                board_list.append(board_dict)
            
            return jsonify({'boards': board_list}), 200
        except Exception as e:
            print(f"Error getting boards: {e}")
            logger.error(traceback.format_exc())
            return jsonify({'error': 'Failed to get boards'}), 500

    @app.route('/api/boards', methods=['POST'])
    @require_auth
    def create_board():
        """Create a new board"""
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({'error': 'Board name is required'}), 400
        
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        # Check board limit
        user_boards = db_manager.get_user_boards(current_user_id)
        if len(user_boards) >= app.config['MAX_BOARDS_PER_USER']:
            return jsonify({'error': 'Maximum number of boards reached'}), 400
        
        board_data = {
            'name': data['name'],
            'description': data.get('description', ''),
            'owner_id': current_user_id,
            'currency': data.get('currency', 'ILS'),
            'timezone': data.get('timezone', 'Asia/Jerusalem'),
            'settings': data.get('settings', {}),
            'board_type': data.get('board_type', 'general')
        }
        
        board = db_manager.create_board(board_data)
        
        # Handle custom categories if provided
        custom_categories = data.get('custom_categories', [])
        if custom_categories and len(custom_categories) > 0:
            print(f"📋 Creating {len(custom_categories)} custom categories for board {board.id}")
            for category_data in custom_categories:
                category_info = {
                    'board_id': board.id,
                    'name': category_data.get('name'),
                    'icon': category_data.get('icon', 'ellipsis-horizontal'),
                    'color': category_data.get('color', '#9370DB'),
                    'image_url': category_data.get('imageUrl'),  # Add image URL support
                    'created_by': current_user_id,
                    'is_default': False
                }
                try:
                    db_manager.create_category(category_info)
                    print(f"✅ Created custom category: {category_info['name']}")
                except Exception as e:
                    print(f"❌ Failed to create custom category {category_info['name']}: {str(e)}")
        
        return jsonify({
            'id': board.id,
            'name': board.name,
            'description': board.description,
            'owner_id': board.owner_id,
            'created_at': board.created_at,
            'currency': board.currency,
            'timezone': board.timezone,
            'board_type': board.board_type
        }), 201

    @app.route('/api/boards/<board_id>', methods=['GET'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def get_board(board_id):
        """Get board details"""
        # לא צריך לבדוק איפוס כאן - זה יקרה בget_budget_status או calculate_board_total_expenses
        
        board = db_manager.get_board_by_id(board_id)
        if not board:
            return jsonify({'error': 'Board not found'}), 404
        
        members = db_manager.get_board_members(board_id)
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        return jsonify({
            'id': board.id,
            'name': board.name,
            'description': board.description,
            'owner_id': board.owner_id,
            'created_at': board.created_at,
            'updated_at': board.updated_at,
            'currency': board.currency,
            'timezone': board.timezone,
            'board_type': board.board_type,
            'settings': board.settings,
            'budget_amount': board.budget_amount,
            'budget_alerts': board.budget_alerts or [],
            'budget_auto_reset': board.budget_auto_reset,
            'budget_reset_day': board.budget_reset_day,
            'budget_reset_time': format_time_field(board.budget_reset_time),
            'budget_last_reset': format_datetime_field(board.budget_last_reset),
            'members': [
                {
                    'id': member.id,
                    'user_id': member.user_id,
                    'role': member.role,
                    'joined_at': member.joined_at,
                    'permissions': member.permissions
                } for member in members
            ],
            'user_role': db_manager.get_user_role_in_board(current_user_id, board_id)
        }), 200

    @app.route('/api/boards/<board_id>', methods=['PUT'])
    @require_auth
    @require_board_admin
    def update_board(board_id):
        """Update board details"""
        data = request.get_json()
        print(f"💰 Update board API called with data: {data}")
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        board = db_manager.get_board_by_id(board_id)
        if not board:
            return jsonify({'error': 'Board not found'}), 404
        
        update_data = {}
        allowed_fields = ['name', 'description', 'currency', 'timezone', 'settings', 'budget_amount', 'budget_alerts', 'budget_auto_reset', 'budget_reset_day', 'budget_reset_time']
        
        for field in allowed_fields:
            if field in data:
                update_data[field] = data[field]
        
        print(f"💰 Filtered update data: {update_data}")
        
        if update_data:
            success = db_manager.update_board(board_id, update_data)
            if success:
                updated_board = db_manager.get_board_by_id(board_id)
                return jsonify({
                    'id': updated_board.id,
                    'name': updated_board.name,
                    'description': updated_board.description,
                    'owner_id': updated_board.owner_id,
                    'created_at': updated_board.created_at,
                    'updated_at': updated_board.updated_at,
                    'currency': updated_board.currency,
                    'timezone': updated_board.timezone,
                    'settings': updated_board.settings,
                    'budget_amount': updated_board.budget_amount,
                    'budget_alerts': updated_board.budget_alerts or [],
                    'budget_auto_reset': updated_board.budget_auto_reset,
                    'budget_reset_day': updated_board.budget_reset_day,
                    'budget_reset_time': format_time_field(updated_board.budget_reset_time),
                    'budget_last_reset': format_datetime_field(updated_board.budget_last_reset)
                }), 200
        
        return jsonify({'error': 'No valid fields to update'}), 400

    @app.route('/api/boards/<board_id>', methods=['DELETE'])
    @require_auth
    @require_board_owner
    def delete_board(board_id):
        """Delete board"""
        board = db_manager.get_board_by_id(board_id)
        if not board:
            return jsonify({'error': 'Board not found'}), 404
        
        success = db_manager.delete_board(board_id)
        if success:
            return jsonify({'message': 'Board deleted successfully'}), 200
        else:
            return jsonify({'error': 'Failed to delete board'}), 500

    # Board member routes
    @app.route('/api/boards/<board_id>/members', methods=['GET'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def get_board_members(board_id):
        """Get board members"""
        members = db_manager.get_board_members(board_id)
        print(f"🔍 Board {board_id} has {len(members)} members:")
        for member in members:
            user = db_manager.get_user_by_id(member.user_id)
            if user:
                print(f"🔍 Member: {user.first_name} {user.last_name} (ID: {user.id}, Role: {member.role})")
        
        member_list = []
        for member in members:
            user = db_manager.get_user_by_id(member.user_id)
            if user:
                member_list.append({
                    'id': member.id,
                    'user_id': member.user_id,
                    'role': member.role,
                    'joined_at': member.joined_at,
                    'permissions': member.permissions,
                    'user': {
                        'id': user.id,
                        'email': user.email,
                        'username': user.username,
                        'first_name': user.first_name,
                        'last_name': user.last_name,
                        'avatar_url': user.avatar_url
                    }
                })
        
        return jsonify({'members': member_list}), 200

    @app.route('/api/boards/<board_id>/members', methods=['POST'])
    @require_auth
    @require_board_admin
    def invite_member(board_id):
        """Invite a member to the board or add a virtual member"""
        data = request.get_json()
        
        # Check if this is a virtual member (no email, just names)
        is_virtual = not data.get('email') and data.get('first_name') and data.get('last_name')
        
        if not data or (not data.get('email') and not is_virtual):
            return jsonify({'error': 'Either email or first_name and last_name are required'}), 400
        
        board = db_manager.get_board_by_id(board_id)
        if not board:
            return jsonify({'error': 'Board not found'}), 404
        
        # Check member limit
        current_members = db_manager.get_board_members(board_id)
        if len(current_members) >= app.config['MAX_USERS_PER_BOARD']:
            return jsonify({'error': 'Maximum number of members reached'}), 400
        
        if is_virtual:
            # Create virtual user and add to board
            try:
                virtual_user = db_manager.create_virtual_user(
                    first_name=data['first_name'].strip(),
                    last_name=data['last_name'].strip()
                )
                
                member = db_manager.add_board_member(
                    board_id=board_id,
                    user_id=virtual_user.id,
                    role=data.get('role', UserRole.MEMBER.value),
                    invited_by=auth_manager.get_current_user()['user']['id']
                )
                
                return jsonify({
                    'message': 'Virtual member added successfully',
                    'member_id': member.id,
                    'is_virtual': True
                }), 201
                
            except Exception as e:
                print(f"❌ Error creating virtual member: {e}")
                return jsonify({'error': 'Failed to create virtual member'}), 500
        else:
            # Regular email-based invitation
            # Convert email to lowercase
            data['email'] = data['email'].lower()
            
            # Check if user already exists
            user = db_manager.get_user_by_email(data['email'])
            if user:
                # User exists, add them directly
                member = db_manager.add_board_member(
                    board_id=board_id,
                    user_id=user.id,
                    role=data.get('role', UserRole.MEMBER.value),
                    invited_by=auth_manager.get_current_user()['user']['id']
                )
                return jsonify({
                    'message': 'Member added successfully',
                    'member_id': member.id
                }), 201
            else:
                # User doesn't exist, create invitation
                current_user_info = auth_manager.get_current_user()['user']
                inviter = db_manager.get_user_by_id(current_user_info['id'])
                inviter_name = f"{inviter.first_name} {inviter.last_name}" if inviter else "משתמש לא ידוע"
                
                invitation_data = {
                    'board_id': board_id,
                    'email': data['email'],
                    'invited_by': current_user_info['id'],
                    'role': data.get('role', UserRole.MEMBER.value),
                    'expires_at': (datetime.now() + timedelta(days=7)).isoformat()
                }
                
                invitation = db_manager.create_invitation(invitation_data)
                
                # Send invitation email
                try:
                    email_sent = auth_manager.send_board_invitation_email(
                        email=data['email'],
                        inviter_name=inviter_name,
                        board_name=board.name,
                        invitation_token=invitation.token,
                        app_config=app.config
                    )
                    
                    if email_sent:
                        print(f"✅ Invitation email sent successfully to {data['email']} for board '{board.name}'")
                    else:
                        print(f"⚠️ Failed to send invitation email to {data['email']}, but invitation was created")
                        
                except Exception as e:
                    print(f"❌ Error sending invitation email: {e}")
                    # Continue anyway - invitation was created
                
                return jsonify({
                    'message': 'Invitation sent successfully',
                    'invitation_id': invitation.id
                }), 201

    @app.route('/api/boards/<board_id>/members/<user_id>', methods=['DELETE'])
    @require_auth
    @require_board_admin
    def remove_member(board_id, user_id):
        """Remove a member from the board"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        # Check if current user can modify this member
        if not db_manager.can_modify_member(current_user_id, board_id, user_id):
            if user_id == current_user_id:
                return jsonify({'error': 'Cannot remove yourself from the board'}), 400
            elif db_manager.is_board_creator(user_id, board_id):
                return jsonify({'error': 'Cannot remove the board creator'}), 403
            else:
                return jsonify({'error': 'Insufficient permissions to remove this member'}), 403
        
        if db_manager.remove_board_member(board_id, user_id):
            return jsonify({'message': 'Member removed successfully'}), 200
        else:
            return jsonify({'error': 'Failed to remove member'}), 500

    @app.route('/api/boards/<board_id>/members/<user_id>/role', methods=['PUT'])
    @require_auth
    @require_board_admin
    def update_member_role(board_id, user_id):
        """Update a member's role in the board"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        data = request.get_json()
        
        if not data or not data.get('role'):
            return jsonify({'error': 'Role is required'}), 400
        
        new_role = data['role']
        
        # Validate role
        valid_roles = [UserRole.OWNER.value, UserRole.ADMIN.value, UserRole.MEMBER.value, UserRole.VIEWER.value]
        if new_role not in valid_roles:
            return jsonify({'error': 'Invalid role'}), 400
        
        # Check if current user can modify this member
        if not db_manager.can_modify_member(current_user_id, board_id, user_id):
            if user_id == current_user_id:
                return jsonify({'error': 'Cannot modify your own role'}), 400
            elif db_manager.is_board_creator(user_id, board_id):
                return jsonify({'error': 'Cannot modify the board creator\'s role'}), 403
            else:
                return jsonify({'error': 'Insufficient permissions to modify this member\'s role'}), 403
        
        # Get the member record
        from postgres_models import BoardMember
        member = BoardMember.query.filter_by(board_id=board_id, user_id=user_id, is_active=True).first()
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        # Update the role
        if db_manager.update_board_member_role(member.id, new_role):
            return jsonify({'message': 'Member role updated successfully'}), 200
        else:
            return jsonify({'error': 'Failed to update member role'}), 500

    @app.route('/api/boards/<board_id>/set-default', methods=['PUT'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def set_default_board(board_id):
        """Set a board as the user's default board"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        success = db_manager.set_default_board(current_user_id, board_id)
        
        if success:
            return jsonify({'message': 'Default board set successfully'}), 200
        else:
            return jsonify({'error': 'Failed to set default board'}), 500

    @app.route('/api/boards/clear-default', methods=['PUT'])
    @require_auth
    def clear_default_board():
        """Clear the user's default board"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        success = db_manager.clear_default_board(current_user_id)
        
        if success:
            return jsonify({'message': 'Default board cleared successfully'}), 200
        else:
            return jsonify({'error': 'Failed to clear default board'}), 500

    @app.route('/api/boards/<board_id>/export-expenses', methods=['POST'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def export_board_expenses(board_id):
        """Export board expenses to an Excel report"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            
            current_user_id = auth_manager.get_current_user()['user']['id']
            
            # Get date filters from request
            data = request.get_json() or {}
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            
            # Get board details
            board = db_manager.get_board_by_id(board_id)
            if not board:
                return jsonify({'error': 'Board not found'}), 404
            
            # Check if this is a work management board
            is_work_management = board.board_type == 'work_management'
            
            # Get all expenses for the board
            expenses = db_manager.get_board_expenses(board_id)
            if not expenses:
                return jsonify({'error': 'No expenses found for this board'}), 404
            
            # Filter expenses by date range if provided
            if start_date or end_date:
                filtered_expenses = []
                for expense in expenses:
                    expense_date = normalize_expense_date(expense.date)
                    
                    # Check start date
                    if start_date:
                        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                        if expense_date < start_dt:
                            continue
                    
                    # Check end date
                    if end_date:
                        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                        # Add one day to end_date to include the entire end day
                        end_dt = end_dt.replace(hour=23, minute=59, second=59)
                        if expense_date > end_dt:
                            continue
                    
                    filtered_expenses.append(expense)
                
                expenses = filtered_expenses
                
                if not expenses:
                    return jsonify({'error': 'No expenses found for the selected date range'}), 404
            
            # Get board members for names
            members = db_manager.get_board_members(board_id)
            member_names = {}
            for member in members:
                user = db_manager.get_user_by_id(member.user_id)
                if user:
                    member_names[member.user_id] = f"{user.first_name} {user.last_name}"
            
            # Create Excel workbook
            wb = Workbook()
            
            # Main expenses sheet
            ws_expenses = wb.active
            ws_expenses.title = "עבודות" if is_work_management else "הוצאות"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Board info header
            ws_expenses['A1'] = "שם לוח:"
            ws_expenses['B1'] = board.name
            ws_expenses['A2'] = "תאריך ייצוא:"
            ws_expenses['B2'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            ws_expenses['A3'] = f"סה\"כ {'עבודות' if is_work_management else 'הוצאות'}:"
            ws_expenses['B3'] = len(expenses)
            ws_expenses['A4'] = "סה\"כ סכום:"
            ws_expenses['B4'] = sum(expense.amount for expense in expenses)
            ws_expenses['A5'] = "מטבע:"
            ws_expenses['B5'] = board.currency
            
            # Add date range info if filtering was applied
            header_row_start = 7
            if start_date or end_date:
                ws_expenses['A6'] = "טווח תאריכים:"
                date_range = ""
                if start_date and end_date:
                    start_formatted = datetime.fromisoformat(start_date.replace('Z', '+00:00')).strftime('%d/%m/%Y')
                    end_formatted = datetime.fromisoformat(end_date.replace('Z', '+00:00')).strftime('%d/%m/%Y')
                    date_range = f"{start_formatted} - {end_formatted}"
                elif start_date:
                    start_formatted = datetime.fromisoformat(start_date.replace('Z', '+00:00')).strftime('%d/%m/%Y')
                    date_range = f"מ-{start_formatted}"
                elif end_date:
                    end_formatted = datetime.fromisoformat(end_date.replace('Z', '+00:00')).strftime('%d/%m/%Y')
                    date_range = f"עד {end_formatted}"
                ws_expenses['B6'] = date_range
                ws_expenses['A6'].font = Font(bold=True)
                header_row_start = 8
            
            # Style the header info
            header_info_rows = 6 if (start_date or end_date) else 5
            for row in range(1, header_info_rows + 1):
                ws_expenses[f'A{row}'].font = Font(bold=True)
            
            # Expenses table header
            if is_work_management:
                headers = ["תיאור", "קטגוריה", "סכום", "בוצע על ידי", "תאריך עבודה", "תאריך יצירה", "תגיות", "הוצאה חוזרת", "תדירות"]
            else:
                headers = ["תיאור", "קטגוריה", "סכום", "שולם על ידי", "תאריך הוצאה", "תאריך יצירה", "תגיות", "הוצאה חוזרת", "תדירות"]
            header_row = header_row_start
            
            for col, header in enumerate(headers, 1):
                cell = ws_expenses.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Add expense data
            current_row = header_row + 1
            for expense in expenses:
                if is_work_management and expense.work_data:
                    # For work management, create detailed rows for each work item
                    work_data = expense.work_data
                    for work_item in work_data.get('workItems', []):
                        description = f"{work_item.get('category', '')} - {work_item.get('description', '')}" if work_item.get('description') else work_item.get('category', '')
                        category = work_item.get('category', '')
                        amount = work_item.get('price', 0)
                        paid_by = member_names.get(expense.paid_by, 'Unknown')
                        
                        # Format expense date (when the work happened)
                        date = datetime.fromisoformat(expense.date.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
                        
                        # Format created date (when the record was added to database)
                        created_date = datetime.fromisoformat(expense.created_at.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
                        
                        tags = '; '.join(expense.tags) if expense.tags else ''
                        is_recurring = 'כן' if expense.is_recurring else 'לא'
                        frequency = expense.frequency if expense.is_recurring else ''
                        
                        # Add work-specific info
                        if work_item.get('hours'):
                            description += f" ({work_item.get('hours')} שעות)"
                        
                        row_data = [description, category, amount, paid_by, date, created_date, tags, is_recurring, frequency]
                        
                        for col, value in enumerate(row_data, 1):
                            cell = ws_expenses.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if col == 3:  # Amount column (סכום)
                                cell.alignment = Alignment(horizontal="right")
                            elif col in [5, 6]:  # Date columns (תאריך עבודה, תאריך יצירה)
                                cell.alignment = Alignment(horizontal="center")
                        
                        current_row += 1
                else:
                    # Regular expense handling
                    description = expense.description or ''
                    category = expense.category or ''
                    amount = expense.amount or 0
                    paid_by = member_names.get(expense.paid_by, 'Unknown')
                    
                    # Format expense date (when the expense happened)
                    date = datetime.fromisoformat(expense.date.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
                    
                    # Format created date (when the record was added to database)
                    created_date = datetime.fromisoformat(expense.created_at.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
                    
                    tags = '; '.join(expense.tags) if expense.tags else ''
                    is_recurring = 'כן' if expense.is_recurring else 'לא'
                    frequency = expense.frequency if expense.is_recurring else ''
                    
                    row_data = [description, category, amount, paid_by, date, created_date, tags, is_recurring, frequency]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws_expenses.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        if col == 3:  # Amount column (סכום)
                            cell.alignment = Alignment(horizontal="right")
                        elif col in [5, 6]:  # Date columns (תאריך הוצאה, תאריך יצירה)
                            cell.alignment = Alignment(horizontal="center")
                    
                    current_row += 1
            
            # Auto-adjust column widths
            for column in ws_expenses.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_expenses.column_dimensions[column_letter].width = adjusted_width
            
            # Category summary sheet
            ws_category = wb.create_sheet("סיכום קטגוריות")
            
            # Category summary header
            ws_category['A1'] = "קטגוריה"
            ws_category['B1'] = f"מספר {'עבודות' if is_work_management else 'הוצאות'}"
            ws_category['C1'] = "סה\"כ סכום"
            
            for col in range(1, 4):
                cell = ws_category.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Calculate category summary
            category_summary = {}
            for expense in expenses:
                if expense.category not in category_summary:
                    category_summary[expense.category] = {'count': 0, 'total': 0}
                category_summary[expense.category]['count'] += 1
                category_summary[expense.category]['total'] += expense.amount
            
            # Add category data
            row = 2
            for category, summary in category_summary.items():
                ws_category.cell(row=row, column=1, value=category).border = border
                ws_category.cell(row=row, column=2, value=summary['count']).border = border
                ws_category.cell(row=row, column=3, value=summary['total']).border = border
                row += 1
            
            # Auto-adjust category sheet columns
            for column in ws_category.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws_category.column_dimensions[column_letter].width = adjusted_width
            
            # Monthly summary sheet
            ws_monthly = wb.create_sheet("סיכום חודשי")
            
            # Monthly summary header
            ws_monthly['A1'] = "חודש"
            ws_monthly['B1'] = f"מספר {'עבודות' if is_work_management else 'הוצאות'}"
            ws_monthly['C1'] = "סה\"כ סכום"
            
            for col in range(1, 4):
                cell = ws_monthly.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Calculate monthly summary
            monthly_summary = {}
            for expense in expenses:
                expense_date = normalize_expense_date(expense.date)
                month_key = expense_date.strftime('%Y-%m')
                if month_key not in monthly_summary:
                    monthly_summary[month_key] = {'count': 0, 'total': 0}
                monthly_summary[month_key]['count'] += 1
                monthly_summary[month_key]['total'] += expense.amount
            
            # Add monthly data
            row = 2
            for month, summary in sorted(monthly_summary.items()):
                month_name = datetime.strptime(month + '-01', '%Y-%m-%d').strftime('%B %Y')
                ws_monthly.cell(row=row, column=1, value=month_name).border = border
                ws_monthly.cell(row=row, column=2, value=summary['count']).border = border
                ws_monthly.cell(row=row, column=3, value=summary['total']).border = border
                row += 1
            
            # Auto-adjust monthly sheet columns
            for column in ws_monthly.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws_monthly.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Create filename (English only for compatibility)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            # Use only ASCII characters in filename to avoid encoding issues
            import re
            board_name_ascii = re.sub(r'[^\x00-\x7F]+', '', board.name)  # Remove non-ASCII
            board_name_clean = re.sub(r'[^\w\-_]', '_', board_name_ascii)  # Keep only word chars, hyphens, underscores
            board_name_clean = re.sub(r'_+', '_', board_name_clean).strip('_')  # Remove multiple underscores
            if not board_name_clean or len(board_name_clean) < 2:  # If no valid name
                board_name_clean = "board"
            filename = f"{'works' if is_work_management else 'expenses'}_{board_name_clean}_{timestamp}.xlsx"
            
            # Return Excel file as download
            response = Response(
                excel_buffer.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename={filename}',  # No quotes to avoid encoding issues
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8'
                }
            )
            
            print(f"✅ Excel export successful for board '{board.name}' - {len(expenses)} {'works' if is_work_management else 'expenses'}")
            return response
            
        except Exception as e:
            print(f"Error exporting {'works' if is_work_management else 'expenses'}: {e}")
            logger.error(traceback.format_exc())
            return jsonify({'error': f'Failed to export {'works' if is_work_management else 'expenses'}'}), 500

    # Expense routes
    @app.route('/api/boards/<board_id>/expenses', methods=['GET'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def get_board_expenses(board_id):
        """Get expenses for a board"""
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        
        expenses = db_manager.get_board_expenses(board_id, month, year)
        
        expense_list = []
        for expense in expenses:
            expense_dict = {
                'id': expense.id,
                'board_id': expense.board_id,
                'amount': expense.amount,
                'category': expense.category,
                'description': expense.description,
                'paid_by': expense.paid_by,
                'date': expense.date,
                'created_by': expense.created_by,
                'created_at': expense.created_at,
                'updated_at': expense.updated_at,
                'is_recurring': expense.is_recurring,
                'frequency': expense.frequency,
                'start_date': expense.start_date,
                'end_date': expense.end_date,
                'tags': expense.tags or [],
                'has_image': bool(expense.receipt_url)  # Boolean flag instead of URL
            }
            expense_list.append(expense_dict)
        
        return jsonify({'expenses': expense_list}), 200

    @app.route('/api/boards/<board_id>/expenses', methods=['POST'])
    @require_auth
    @require_board_access(BoardPermission.WRITE.value)
    def create_expense(board_id):
        """Create a new expense"""
        data = request.get_json()
        if not data or not data.get('amount') or not data.get('category'):
            return jsonify({'error': 'Amount and category are required'}), 400
        
        current_user_id = auth_manager.get_current_user()['user']['id']
        print(f"🔍 Creating expense - User: {current_user_id}, Board: {board_id}, Amount: {data['amount']}")
        
        expense_data = {
            'board_id': board_id,
            'amount': data['amount'],
            'category': data['category'],
            'description': data.get('description', ''),
            'paid_by': data.get('paid_by', current_user_id),
            'date': data.get('date', datetime.now().isoformat()),
            'created_by': current_user_id,
            'is_recurring': data.get('is_recurring', False),
            'frequency': data.get('frequency', 'monthly'),
            'start_date': data.get('start_date'),
            'end_date': data.get('end_date'),
            'receipt_url': data.get('image_url'),  # Use image_url from frontend
            'tags': data.get('tags', []),
            'work_data': data.get('work_data')  # Add work management data
        }
        
        print(f"🔍 Expense data - Paid by: {expense_data['paid_by']}")
        
        expense = db_manager.create_expense(expense_data)
        print(f"🔍 Expense created with ID: {expense.id}")
        
        # Create debts if there are multiple members
        members = db_manager.get_board_members(board_id)
        print(f"🔍 Board has {len(members)} members")
        for member in members:
            user = db_manager.get_user_by_id(member.user_id)
            print(f"🔍 Member: {user.first_name} {user.last_name} (ID: {member.user_id})")
        
        if len(members) > 1:
            print(f"💰 Processing debts for expense...")
            process_expense_debts(expense, members)
        else:
            print(f"💰 Only one member - no debts to process")
        
        # Create notifications for all board members except the creator (original functionality)
        db_manager.create_expense_notification(expense, members, 'expense_added')
        
        # Additionally, if someone added an expense for another person, send special notification
        if expense.paid_by != expense.created_by:
            # Find the member who the expense was added for
            payer_member = next((m for m in members if m.user_id == expense.paid_by), None)
            if payer_member:
                # Send additional notification to the person who the expense was added for
                db_manager.create_expense_notification(expense, [payer_member], 'expense_added_for_you')
        
        return jsonify({
            'id': expense.id,
            'board_id': expense.board_id,
            'amount': expense.amount,
            'category': expense.category,
            'description': expense.description,
            'paid_by': expense.paid_by,
            'date': expense.date,
            'created_by': expense.created_by,
            'created_at': expense.created_at,
            'is_recurring': expense.is_recurring,
            'frequency': expense.frequency,
            'tags': expense.tags or [],
            'has_image': bool(expense.receipt_url)  # Boolean flag instead of URL
        }), 201

    @app.route('/api/boards/<board_id>/expenses/<expense_id>', methods=['PUT'])
    @require_auth
    @require_board_access(BoardPermission.WRITE.value)
    def update_expense(board_id, expense_id):
        """Update an expense"""
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Get the expense to check if it exists and belongs to the board
        expenses = db_manager.get_board_expenses(board_id)
        expense = next((e for e in expenses if e.id == expense_id), None)
        
        if not expense:
            return jsonify({'error': 'Expense not found'}), 404
        
        # Check if there are critical updates that would affect paid debts
        if 'amount' in data or 'paid_by' in data:
            # Check if there are any paid debts associated with this expense
            from postgres_models import Debt
            paid_debts = Debt.query.filter_by(expense_id=expense_id, is_paid=True).all()
            
            if paid_debts:
                return jsonify({'error': 'Cannot modify amount or payer - some payments have already been made. Contact support if you need to modify this expense.'}), 400
        
        update_data = {}
        allowed_fields = ['amount', 'category', 'description', 'paid_by', 'date', 'is_recurring', 'frequency', 'start_date', 'end_date', 'tags']
        
        for field in allowed_fields:
            if field in data:
                update_data[field] = data[field]
        
        # Handle image_url from frontend, map to receipt_url in database
        if 'image_url' in data:
            update_data['receipt_url'] = data['image_url']
        
        if update_data:
            success = db_manager.update_expense(expense_id, update_data)
            if success:
                # Update debts only if amount or paid_by changed AND there are no paid debts
                if 'amount' in update_data or 'paid_by' in update_data:
                    print(f"🔍 Updating debts for expense {expense_id} - amount or payer changed")
                    members = db_manager.get_board_members(board_id)
                    if len(members) > 1:
                        # Remove only UNPAID debts and create new ones
                        print(f"🔍 Removing old UNPAID debts for expense {expense_id}")
                        from postgres_models import Debt
                        Debt.query.filter_by(expense_id=expense_id, is_paid=False).delete()
                        postgres_db.session.commit()
                        updated_expense = db_manager.get_board_expenses(board_id)
                        updated_expense = next((e for e in updated_expense if e.id == expense_id), None)
                        if updated_expense:
                            print(f"🔍 Creating new debts for updated expense")
                            create_debts_for_expense(updated_expense, members)
                
                # Create notifications for expense update
                members = db_manager.get_board_members(board_id)
                updated_expense = db_manager.get_board_expenses(board_id)
                updated_expense = next((e for e in updated_expense if e.id == expense_id), None)
                if updated_expense:
                    db_manager.create_expense_notification(updated_expense, members, 'expense_updated')
                

                
                return jsonify({'message': 'Expense updated successfully'}), 200
        
        return jsonify({'error': 'No valid fields to update'}), 400

    @app.route('/api/boards/<board_id>/expenses/<expense_id>', methods=['DELETE'])
    @require_auth
    @require_board_access(BoardPermission.WRITE.value)
    def delete_expense(board_id, expense_id):
        """Delete an expense"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        # Get the expense before deleting for notifications
        expenses = db_manager.get_board_expenses(board_id)
        expense = next((e for e in expenses if e.id == expense_id), None)
        
        if not expense:
            return jsonify({'error': 'Expense not found'}), 404
        
        # Check if user is the creator of the expense
        if expense.created_by != current_user_id:
            return jsonify({'error': 'You can only delete expenses you created'}), 403
        
        # Check if there are any PAID debts associated with this expense
        # Only paid debts prevent deletion - unpaid debts can be deleted
        from postgres_models import Debt
        paid_debts = Debt.query.filter_by(expense_id=expense_id, is_paid=True).all()
        
        if paid_debts:
            return jsonify({'error': 'Cannot delete expense - some payments have already been made. Contact support if you need to modify this expense.'}), 400
        
        # Get board members for notifications
        members = db_manager.get_board_members(board_id)
        
        # Create notifications before deleting
        db_manager.create_expense_notification(expense, members, 'expense_deleted')
        
        if db_manager.delete_expense(expense_id):
            return jsonify({'message': 'Expense deleted successfully'}), 200
        return jsonify({'error': 'Failed to delete expense'}), 500

    # File upload endpoints
    @app.route('/api/uploads/expense_images/<filename>', methods=['GET'])
    def serve_expense_image(filename):
        """Serve uploaded expense images"""
        try:
            print(f"📁 Serving image: {filename}")
            
            # Security check - ensure filename is safe
            if '..' in filename or '/' in filename or '\\' in filename:
                return jsonify({'error': 'Invalid filename'}), 400
            
            # Serve the image based on storage method
            if app.config.get('UPLOAD_METHOD') == 'b2' and app.b2_service:
                # Get from B2
                file_path = f'expense_images/{filename}'
                try:
                    success, file_content, error = app.b2_service.download_file(file_path)
                    
                    if not success:
                        print(f"❌ B2 file not found: {file_path}")
                        return jsonify({'error': 'Image not found'}), 404
                    
                    # Return the image with proper headers
                    return Response(
                        file_content,
                        mimetype='image/jpeg',
                        headers={'Content-Disposition': f'inline; filename="{filename}"'}
                    )
                        
                except Exception as e:
                    print(f"❌ Error getting B2 image {filename}: {str(e)}")
                    return jsonify({'error': 'Error loading image'}), 500
            else:
                # Get from local storage
                try:
                    file_path = os.path.join(app.config['EXPENSE_IMAGES_FOLDER'], filename)
                    
                    if not os.path.exists(file_path):
                        print(f"❌ Local file not found: {file_path}")
                        return jsonify({'error': 'Image not found'}), 404
                    
                    return send_from_directory(
                        app.config['EXPENSE_IMAGES_FOLDER'], 
                        filename,
                        mimetype='image/jpeg'
                    )
                        
                except Exception as e:
                    print(f"❌ Error reading local image {filename}: {str(e)}")
                    return jsonify({'error': 'Image not found'}), 404
                    
        except Exception as e:
            print(f"❌ Error serving image {filename}: {str(e)}")
            return jsonify({'error': 'Internal server error'}), 500

    @app.route('/api/upload/expense-image', methods=['POST'])
    @require_auth
    def upload_expense_image_endpoint():
        """Upload an image for an expense"""
        if 'image' not in request.files:
            return jsonify({'error': 'No image file provided'}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        image_url = upload_expense_image(file, current_user_id)
        if image_url:
            return jsonify({'image_url': image_url}), 200
        else:
            return jsonify({'error': 'Invalid file type. Allowed types: png, jpg, jpeg, gif, webp'}), 400

    @app.route('/api/expenses/<expense_id>/image', methods=['POST'])
    @require_auth
    def get_expense_image(expense_id):
        """Get image for a specific expense with proper authorization checks"""
        print(f"🔍 === IMAGE ENDPOINT CALLED ===")
        print(f"🔍 Request URL: {request.url}")
        print(f"🔍 Request method: {request.method}")
        print(f"🔍 Expense ID: {expense_id}")
        print(f"🔍 Request headers: {dict(request.headers)}")
        
        # Get current user ID here where auth context is available
        try:
            current_user_result = auth_manager.get_current_user()
            print(f"🔍 Auth result: {current_user_result}")
            
            if not current_user_result.get('valid'):
                return jsonify({'error': 'Authentication failed'}), 401
            
            current_user_id = current_user_result['user']['id']
            print(f"🔍 Current user ID: {current_user_id}")
            
            return get_expense_image_actual(expense_id, current_user_id)
        except Exception as e:
            traceback.print_exc()
            print(f"🔍 Auth error: {e}")
            return jsonify({'error': 'Authentication failed'}), 401

    @app.route('/api/categories/<category_id>/image', methods=['POST'])
    @require_auth
    def get_category_image(category_id):
        """Get image for a specific category with proper authorization checks"""
        print(f"🖼️ === CATEGORY IMAGE ENDPOINT CALLED ===")
        print(f"🖼️ Request URL: {request.url}")
        print(f"🖼️ Request method: {request.method}")
        print(f"🖼️ Category ID: {category_id}")
        print(f"🖼️ Request headers: {dict(request.headers)}")
        
        # Get current user ID here where auth context is available
        try:
            current_user_result = auth_manager.get_current_user()
            print(f"🖼️ Auth result: {current_user_result}")
            
            if not current_user_result.get('valid'):
                return jsonify({'error': 'Authentication failed'}), 401
            
            current_user_id = current_user_result['user']['id']
            print(f"🖼️ Current user ID: {current_user_id}")
            
            return get_category_image_actual(category_id, current_user_id)
        except Exception as e:
            traceback.print_exc()
            print(f"🖼️ Auth error: {e}")
            return jsonify({'error': 'Authentication failed'}), 401
    
    def get_expense_image_actual(expense_id, current_user_id):
        """Get image for a specific expense with proper authorization checks"""
        print(f"🔍 === IMAGE ENDPOINT ACTUAL CALLED ===")
        
        try:
            print(f"🔍 Image endpoint called for expense_id: {expense_id}")
            print(f"🔍 Current user ID: {current_user_id}")
            
            # Always return base64 data for React Native (since we're using POST)
            return_base64 = True
            print(f"🔍 Return base64: {return_base64}")
            
            # Find the expense and check if user has access
            expense = None
            user_boards = db_manager.get_user_boards(current_user_id)
            print(f"🔍 User has access to {len(user_boards)} boards")
            
            for board in user_boards:
                board_expenses = db_manager.get_board_expenses(board.id)
                print(f"🔍 Board {board.name} has {len(board_expenses)} expenses")
                expense = next((e for e in board_expenses if e.id == expense_id), None)
                if expense:
                    print(f"🔍 Found expense in board {board.name}")
                    break
            
            if not expense:
                print(f"🔍 Expense {expense_id} not found or access denied")
                return jsonify({'error': 'Expense not found or access denied'}), 404
            
            print(f"🔍 Expense receipt_url: {expense.receipt_url}")
            if not expense.receipt_url:
                print(f"🔍 No image found for expense {expense_id}")
                return jsonify({'error': 'No image found for this expense'}), 404
            
            # Extract filename from the receipt_url
            # Format: /api/uploads/expense_images/filename.jpg
            if expense.receipt_url.startswith('/api/uploads/expense_images/'):
                filename = expense.receipt_url.split('/')[-1]
            else:
                return jsonify({'error': 'Invalid image URL format'}), 400
            
            # Get file content from storage
            file_content = None
            
            # Serve the image based on storage method
            if app.config.get('UPLOAD_METHOD') == 'b2' and app.b2_service:
                # Get from B2
                file_path = f'expense_images/{filename}'
                try:
                    logger.info(f'Getting image from B2: {file_path}')
                    success, file_content, error = app.b2_service.download_file(file_path)
                    
                    if not success:
                        current_app.logger.error(f"Failed to download B2 file {file_path}: {error}")
                        return jsonify({'error': 'Image not found'}), 404
                        
                except Exception as e:
                    current_app.logger.error(f"Error getting B2 image {filename}: {str(e)}")
                    return jsonify({'error': 'Error loading image'}), 500
            else:
                # Get from local storage
                try:
                    file_path = os.path.join(app.config['EXPENSE_IMAGES_FOLDER'], filename)
                    
                    if not os.path.exists(file_path):
                        return jsonify({'error': 'Image not found'}), 404
                    
                    with open(file_path, 'rb') as f:
                        file_content = f.read()
                        
                except Exception as e:
                    current_app.logger.error(f"Error reading local image {filename}: {str(e)}")
                    return jsonify({'error': 'Image not found'}), 404
            
            # Check if we got file content
            if not file_content:
                return jsonify({'error': 'Image not found'}), 404
            
            # Always return as base64 JSON for React Native
            import base64
            base64_data = base64.b64encode(file_content).decode('utf-8')
            
            return jsonify({
                'image': base64_data
            }), 200
                    
        except Exception as e:
            traceback.print_exc()
            current_app.logger.error(f"Error in get_expense_image: {str(e)}")
            return jsonify({'error': 'Internal server error'}), 500

    def get_category_image_actual(category_id, current_user_id):
        """Get image for a specific category with proper authorization checks"""
        print(f"🖼️ === CATEGORY IMAGE ENDPOINT ACTUAL CALLED ===")
        
        try:
            print(f"🖼️ Category image endpoint called for category_id: {category_id}")
            print(f"🖼️ Current user ID: {current_user_id}")
            
            # Always return base64 data for React Native (since we're using POST)
            return_base64 = True
            print(f"🖼️ Return base64: {return_base64}")
            
            # Find the category and check if user has access
            category = None
            user_boards = db_manager.get_user_boards(current_user_id)
            print(f"🖼️ User has access to {len(user_boards)} boards")
            
            for board in user_boards:
                board_categories = db_manager.get_board_categories(board.id)
                print(f"🖼️ Board {board.name} has {len(board_categories)} categories")
                category = next((c for c in board_categories if c.id == category_id), None)
                if category:
                    print(f"🖼️ Found category in board {board.name}")
                    break
            
            if not category:
                print(f"🖼️ Category {category_id} not found or access denied")
                return jsonify({'error': 'Category not found or access denied'}), 404
            
            print(f"🖼️ Category image_url: {category.image_url}")
            if not category.image_url:
                print(f"🖼️ No image found for category {category_id}")
                return jsonify({'error': 'No image found for this category'}), 404
            
            # Extract filename from the image_url
            # Format: /api/uploads/expense_images/filename.jpg
            if category.image_url.startswith('/api/uploads/expense_images/'):
                filename = category.image_url.split('/')[-1]
            else:
                print(f"🖼️ Invalid image URL format: {category.image_url}")
                return jsonify({'error': 'Invalid image URL format'}), 400
            
            # Get file content from storage
            file_content = None
            
            # Serve the image based on storage method
            if app.config.get('UPLOAD_METHOD') == 'b2' and app.b2_service:
                # Get from B2
                file_path = f'expense_images/{filename}'
                try:
                    logger.info(f'🖼️ Getting category image from B2: {file_path}')
                    success, file_content, error = app.b2_service.download_file(file_path)
                    
                    if not success:
                        current_app.logger.error(f"🖼️ Failed to download B2 file {file_path}: {error}")
                        return jsonify({'error': 'Image not found'}), 404
                        
                except Exception as e:
                    current_app.logger.error(f"🖼️ Error getting B2 category image {filename}: {str(e)}")
                    return jsonify({'error': 'Error loading image'}), 500
            else:
                # Get from local storage
                try:
                    file_path = os.path.join(app.config['EXPENSE_IMAGES_FOLDER'], filename)
                    
                    if not os.path.exists(file_path):
                        print(f"🖼️ Local file not found: {file_path}")
                        return jsonify({'error': 'Image not found'}), 404
                    
                    with open(file_path, 'rb') as f:
                        file_content = f.read()
                        
                except Exception as e:
                    current_app.logger.error(f"🖼️ Error reading local category image {filename}: {str(e)}")
                    return jsonify({'error': 'Image not found'}), 404
            
            # Check if we got file content
            if not file_content:
                print(f"🖼️ No file content for category image: {filename}")
                return jsonify({'error': 'Image not found'}), 404
            
            # Always return as base64 JSON for React Native
            import base64
            base64_data = base64.b64encode(file_content).decode('utf-8')
            
            print(f"🖼️ Successfully returning base64 image data for category {category_id}")
            return jsonify({
                'image': base64_data
            }), 200
                    
        except Exception as e:
            traceback.print_exc()
            current_app.logger.error(f"🖼️ Error in get_category_image: {str(e)}")
            return jsonify({'error': 'Internal server error'}), 500

    # Debt routes
    @app.route('/api/boards/<board_id>/debts', methods=['GET'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def get_board_debts(board_id):
        """Get debts for a board"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        debts = db_manager.get_user_debts(current_user_id, board_id)
        
        debt_list = []
        for debt in debts:
            debt_list.append({
                'id': debt.id,
                'board_id': debt.board_id,
                'expense_id': debt.expense_id,
                'from_user_id': debt.from_user_id,
                'to_user_id': debt.to_user_id,
                'amount': debt.amount,
                'description': debt.description,
                'is_paid': debt.is_paid,
                'paid_at': debt.paid_at,
                'created_at': debt.created_at
            })
        
        return jsonify({'debts': debt_list}), 200

    @app.route('/api/boards/<board_id>/debts/<debt_id>/mark-paid', methods=['PUT'])
    @require_auth
    @require_board_access(BoardPermission.WRITE.value)
    def mark_debt_paid(board_id, debt_id):
        """Mark a debt as paid"""
        success = db_manager.mark_debt_as_paid(debt_id)
        if success:
            return jsonify({'message': 'Debt marked as paid'}), 200
        else:
            return jsonify({'error': 'Failed to mark debt as paid'}), 500

    @app.route('/api/debts/process-partial-payment', methods=['POST'])
    @require_auth
    def process_partial_payment():
        """Process partial payment between users"""
        try:
            data = request.get_json()
            if not data:
                return jsonify({'error': 'No data provided'}), 400
            
            from_user_id = data.get('from_user_id')
            payment_amount = data.get('payment_amount')
            board_ids = data.get('board_ids', [])
            
            if not from_user_id or not payment_amount:
                return jsonify({'error': 'from_user_id and payment_amount are required'}), 400
            
            if payment_amount <= 0:
                return jsonify({'error': 'Payment amount must be positive'}), 400
            
            current_user_id = auth_manager.get_current_user()['user']['id']
            
            # Verify that the current user has access to the boards involved
            user_boards = db_manager.get_user_boards(current_user_id)
            user_board_ids = [board.id for board in user_boards]
            
            if board_ids:
                # Check that all requested board_ids are accessible to the user
                for board_id in board_ids:
                    if board_id not in user_board_ids:
                        return jsonify({'error': f'Access denied to board {board_id}'}), 403
            else:
                # If no specific boards, use all user's boards
                board_ids = user_board_ids
            
            # Process the partial payment
            result = db_manager.process_partial_payment_for_user(
                from_user_id=from_user_id,
                to_user_id=current_user_id,  # Current user is the one receiving payment
                payment_amount=payment_amount,
                board_ids=board_ids
            )
            
            if result['success']:
                return jsonify({
                    'message': 'Partial payment processed successfully',
                    'debts_closed': result['debts_closed'],
                    'debts_updated': result['debts_updated'],
                    'total_processed': result['total_processed']
                }), 200
            else:
                return jsonify({'error': result['error']}), 400
                
        except Exception as e:
            print(f"Error processing partial payment: {e}")
            return jsonify({'error': 'Failed to process partial payment'}), 500

    @app.route('/api/debts/auto-offset', methods=['POST'])
    @require_auth
    def auto_offset_debts():
        """Automatically offset mutual debts between users"""
        try:
            data = request.get_json() or {}
            board_ids = data.get('board_ids', [])
            
            current_user_id = auth_manager.get_current_user()['user']['id']
            
            # Get user's boards
            user_boards = db_manager.get_user_boards(current_user_id)
            user_board_ids = [board.id for board in user_boards]
            
            if board_ids:
                # Check that all requested board_ids are accessible to the user
                for board_id in board_ids:
                    if board_id not in user_board_ids:
                        return jsonify({'error': f'Access denied to board {board_id}'}), 403
            else:
                # If no specific boards, use all user's boards
                board_ids = user_board_ids
            
            # Process debt offsets
            result = db_manager.auto_offset_mutual_debts(current_user_id, board_ids)
            
            if result['success']:
                return jsonify({
                    'message': 'Debt offsets processed successfully',
                    'offsets_processed': result['offsets_processed'],
                    'total_amount_offset': result['total_amount_offset'],
                    'details': result['details']
                }), 200
            else:
                return jsonify({'error': result['error']}), 400
                
        except Exception as e:
            print(f"Error processing debt offsets: {e}")
            return jsonify({'error': 'Failed to process debt offsets'}), 500

    # Summary and Statistics routes
    @app.route('/api/expenses/summary', methods=['GET'])
    @require_auth
    def get_expenses_summary():
        """Get summary statistics for all user's expenses"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        # Get query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        board_ids = request.args.getlist('board_ids')
        
        # Get user's boards
        user_boards = db_manager.get_user_boards(current_user_id)
        if board_ids:
            user_boards = [board for board in user_boards if board.id in board_ids]
        
        if not user_boards:
            return jsonify({
                'total_amount': 0,
                'total_expenses': 0,
                'expenses_by_category': {},
                'expenses_by_board': {},
                'monthly_trend': {}
            }), 200
        
        # Collect all expenses from user's boards
        all_expenses = []
        expenses_by_category = {}
        expenses_by_board = {}
        monthly_trend = {}
        total_amount = 0
        total_expenses = 0
        
        for board in user_boards:
            board_expenses = db_manager.get_board_expenses(board.id)
            
            # Filter by date if provided
            if start_date or end_date:
                filtered_expenses = []
                for expense in board_expenses:
                    expense_date = normalize_expense_date(expense.date)
                    
                    if start_date:
                        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                        if expense_date < start_dt:
                            continue
                    if end_date:
                        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                        if expense_date > end_dt:
                            continue
                    filtered_expenses.append(expense)
                board_expenses = filtered_expenses
            
            # Aggregate statistics
            board_total = sum(expense.amount for expense in board_expenses)
            expenses_by_board[board.name] = board_total
            total_amount += board_total
            total_expenses += len(board_expenses)
            
            for expense in board_expenses:
                all_expenses.append(expense)
                
                # Category aggregation
                if expense.category not in expenses_by_category:
                    expenses_by_category[expense.category] = 0
                expenses_by_category[expense.category] += expense.amount
                
                # Monthly trend
                expense_date = normalize_expense_date(expense.date)
                month_key = expense_date.strftime('%Y-%m')
                if month_key not in monthly_trend:
                    monthly_trend[month_key] = 0
                monthly_trend[month_key] += expense.amount
        
        return jsonify({
            'total_amount': total_amount,
            'total_expenses': total_expenses,
            'expenses_by_category': expenses_by_category,
            'expenses_by_board': expenses_by_board,
            'monthly_trend': monthly_trend
        }), 200

    @app.route('/api/debts/all', methods=['GET'])
    @require_auth
    def get_all_debts():
        """Get all debts for the current user across all boards with filtering support"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        print(f"🔍 === Loading debts for user: {current_user_id} ===")
        
        # Get query parameters for filtering
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        board_ids = request.args.getlist('board_ids')
        is_paid_param = request.args.get('is_paid')
        
        print(f"🔍 Filters received: start_date={start_date}, end_date={end_date}, board_ids={board_ids}, is_paid={is_paid_param}")
        
        # Get user's boards
        user_boards = db_manager.get_user_boards(current_user_id)
        
        # Filter boards if specific board_ids requested
        if board_ids:
            user_boards = [board for board in user_boards if board.id in board_ids]
        
        print(f"🔍 User has access to {len(user_boards)} boards")
        for board in user_boards:
            print(f"🔍 Board: {board.name} (ID: {board.id})")
        
        all_debts = []
        total_owed = 0
        total_owed_to_me = 0
        total_unpaid = 0
        total_paid = 0
        
        for board in user_boards:
            print(f"🔍 Checking debts in board: {board.name} (ID: {board.id})")
            board_debts = db_manager.get_user_debts(current_user_id, board.id)
            print(f"🔍 Found {len(board_debts)} debts in this board")
            
            for debt in board_debts:
                print(f"🔍 Processing debt: {debt.from_user_id} -> {debt.to_user_id}, amount: {debt.amount}, paid: {debt.is_paid}, paid_amount: {debt.paid_amount}")
                
                # Apply is_paid filter if specified
                if is_paid_param is not None:
                    is_paid_filter = is_paid_param.lower() == 'true'
                    if debt.is_paid != is_paid_filter:
                        print(f"🔍 Skipping debt due to paid filter: required={is_paid_filter}, actual={debt.is_paid}")
                        continue
                    else:
                        print(f"✅ Debt passed paid filter: required={is_paid_filter}, actual={debt.is_paid}")
                
                # Apply date filtering if specified (filter by expense creation date)
                if start_date or end_date:
                    try:
                        # Get the expense to check its date
                        expense = db_manager.get_expense_by_id(debt.expense_id)
                        if expense:
                            expense_date = normalize_expense_date(expense.date)
                            
                            # Check start date
                            if start_date:
                                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                                if expense_date < start_dt:
                                    print(f"🔍 Skipping debt due to start date filter: expense_date={expense_date}, start_date={start_dt}")
                                    continue
                            
                            # Check end date
                            if end_date:
                                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                                if expense_date > end_dt:
                                    print(f"🔍 Skipping debt due to end date filter: expense_date={expense_date}, end_date={end_dt}")
                                    continue
                        else:
                            print(f"🔍 Warning: Could not find expense {debt.expense_id} for debt {debt.id}")
                    except Exception as e:
                        print(f"🔍 Error applying date filter to debt {debt.id}: {e}")
                        # Continue processing if date filtering fails
                
                # Get user names for display
                from_user = db_manager.get_user_by_id(debt.from_user_id)
                to_user = db_manager.get_user_by_id(debt.to_user_id)
                
                from_user_name = f"{from_user.first_name} {from_user.last_name}" if from_user else "משתמש לא ידוע"
                to_user_name = f"{to_user.first_name} {to_user.last_name}" if to_user else "משתמש לא ידוע"
                
                debt_data = {
                    'id': debt.id,
                    'board_id': debt.board_id,
                    'expense_id': debt.expense_id,
                    'from_user_id': debt.from_user_id,
                    'to_user_id': debt.to_user_id,
                    'from_user_name': from_user_name,
                    'to_user_name': to_user_name,
                    'amount': debt.amount,
                    'description': debt.description,
                    'is_paid': debt.is_paid,
                    'paid_at': debt.paid_at,
                    'created_at': debt.created_at,
                    'board_name': board.name,
                    'original_amount': debt.original_amount,
                    'paid_amount': debt.paid_amount
                }
                
                print(f"✅ Adding debt to results: {debt.id}, paid={debt.is_paid}, amount={debt.amount}, paid_amount={debt.paid_amount}")
                all_debts.append(debt_data)
                
                # Calculate summary
                if debt.from_user_id == current_user_id:
                    print(f"🔍 Current user owes {debt.amount} to {to_user_name}")
                    if not debt.is_paid:
                        total_owed += debt.amount
                        total_unpaid += debt.amount
                    else:
                        total_paid += debt.amount
                else:
                    print(f"🔍 {from_user_name} owes {debt.amount} to current user")
                    if not debt.is_paid:
                        total_owed_to_me += debt.amount
        
        print(f"🔍 Final summary - Total owed: {total_owed}, Total owed to me: {total_owed_to_me}")
        print(f"🔍 Total debts found after filtering: {len(all_debts)}")
        
        return jsonify({
            'debts': all_debts,
            'summary': {
                'total_owed': total_owed,
                'total_owed_to_me': total_owed_to_me,
                'total_unpaid': total_unpaid,
                'total_paid': total_paid
            }
        }), 200

    @app.route('/api/expenses/by-period', methods=['GET'])
    @require_auth
    def get_expenses_by_period():
        """Get expenses for a specific period"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        board_ids = request.args.getlist('board_ids')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Start date and end date are required'}), 400
        
        # Get user's boards
        user_boards = db_manager.get_user_boards(current_user_id)
        if board_ids:
            user_boards = [board for board in user_boards if board.id in board_ids]
        
        all_expenses = []
        total_amount = 0
        
        for board in user_boards:
            board_expenses = db_manager.get_board_expenses(board.id)
            
            for expense in board_expenses:
                expense_date = normalize_expense_date(expense.date)
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                
                if start_dt <= expense_date <= end_dt:
                    all_expenses.append({
                        'id': expense.id,
                        'board_id': expense.board_id,
                        'amount': expense.amount,
                        'category': expense.category,
                        'description': expense.description,
                        'paid_by': expense.paid_by,
                        'date': expense.date,
                        'created_by': expense.created_by,
                        'created_at': expense.created_at,
                        'is_recurring': expense.is_recurring,
                        'frequency': expense.frequency,
                        'tags': expense.tags or [],
                        'board_name': board.name
                    })
                    total_amount += expense.amount
        
        # Calculate average per day
        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        days_diff = (end_dt - start_dt).days + 1
        average_per_day = total_amount / days_diff if days_diff > 0 else 0
        
        return jsonify({
            'expenses': all_expenses,
            'summary': {
                'total_amount': total_amount,
                'total_expenses': len(all_expenses),
                'average_per_day': average_per_day
            }
        }), 200

    # Category routes
    @app.route('/api/boards/<board_id>/categories', methods=['GET'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def get_board_categories(board_id):
        """Get categories for a board"""
        categories = db_manager.get_board_categories(board_id)
        
        category_list = []
        for category in categories:
            category_list.append({
                'id': category.id,
                'board_id': category.board_id,
                'name': category.name,
                'icon': category.icon,
                'color': category.color,
                'image_url': category.image_url,
                'created_by': category.created_by,
                'created_at': category.created_at,
                'is_default': category.is_default,
                'is_active': category.is_active
            })
        
        return jsonify({'categories': category_list}), 200

    @app.route('/api/boards/<board_id>/categories', methods=['POST'])
    @require_auth
    @require_board_admin
    def create_category(board_id):
        """Create a new category"""
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({'error': 'Category name is required'}), 400
        
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        category_data = {
            'board_id': board_id,
            'name': data['name'],
            'icon': data.get('icon', 'ellipsis-horizontal'),
            'color': data.get('color', '#9370DB'),
            'image_url': data.get('image_url'),  # Add image URL support
            'created_by': current_user_id,
            'is_default': data.get('is_default', False)
        }
        
        category = db_manager.create_category(category_data)
        
        return jsonify({
            'id': category.id,
            'board_id': category.board_id,
            'name': category.name,
            'icon': category.icon,
            'color': category.color,
            'image_url': category.image_url,  # Include image URL in response
            'created_by': category.created_by,
            'created_at': category.created_at,
            'is_default': category.is_default
        }), 201

    @app.route('/api/boards/<board_id>/categories/update', methods=['PUT'])
    @require_auth
    @require_board_admin
    def update_board_categories(board_id):
        """Update all categories for a board"""
        data = request.get_json()
        if not data or 'categories' not in data:
            return jsonify({'error': 'Categories data is required'}), 400
        
        current_user_id = auth_manager.get_current_user()['user']['id']
        categories = data['categories']
        
        try:
            # First, delete all existing custom categories for this board (keep default ones)
            # Find all custom categories for this board
            from postgres_models import Category
            custom_categories = Category.query.filter_by(board_id=board_id, is_default=False).all()
            
            # Remove each custom category by ID
            for category in custom_categories:
                postgres_db.session.delete(category)
            postgres_db.session.commit()
            
            # Create new categories
            for category_data in categories:
                category_info = {
                    'board_id': board_id,
                    'name': category_data.get('name'),
                    'icon': category_data.get('icon', 'ellipsis-horizontal'),
                    'color': category_data.get('color', '#9370DB'),
                    'image_url': category_data.get('image_url'),  # Add image URL support
                    'created_by': current_user_id,
                    'is_default': False
                }
                try:
                    db_manager.create_category(category_info)
                    print(f"✅ Updated category: {category_info['name']}")
                except Exception as e:
                    print(f"❌ Failed to create category {category_info['name']}: {str(e)}")
            
            return jsonify({'message': 'Categories updated successfully'}), 200
            
        except Exception as e:
            print(f"❌ Error updating board categories: {str(e)}")
            return jsonify({'error': 'Failed to update categories'}), 500

    @app.route('/api/boards/<board_id>/budget/status', methods=['GET'])
    @require_auth
    @require_board_access(BoardPermission.READ.value)
    def get_board_budget_status(board_id):
        """Get budget status for a board including alerts"""
        try:
            # בדיקת איפוס תקרה בתוך get_budget_status
            budget_status = get_budget_status(board_id)
            return jsonify(budget_status), 200
        except Exception as e:
            print(f"Error getting budget status: {e}")
            return jsonify({'error': 'Failed to get budget status'}), 500

    # Helper function to process debts for an expense with smart offsetting
    def process_expense_debts(expense, members):
        """
        Process debts for an expense with smart offsetting logic.
        
        Logic:
        1. Calculate how much each member owes for this expense
        2. For each member, check existing debts with the payer
        3. Offset existing debts where possible
        4. Create new debts only for remaining amounts
        """
        from postgres_models import Debt
        
        print(f"💰 === Processing debts for expense ===")
        print(f"💰 Expense ID: {expense.id}")
        print(f"💰 Expense amount: {expense.amount}")
        print(f"💰 Total members: {len(members)}")
        print(f"💰 Paid by: {expense.paid_by}")
        
        payer_id = expense.paid_by
        other_members = [member for member in members if member.user_id != payer_id]
        
        if not other_members:
            print("💰 No other members - no debts to process")
            return
        
        # Calculate how much each person owes
        total_people = len(members)
        amount_per_person = expense.amount / total_people
        print(f"💰 Amount per person: {amount_per_person}")
        
        debts_processed = []
        
        for member in other_members:
            member_id = member.user_id
            member_user = db_manager.get_user_by_id(member_id)
            member_name = f"{member_user.first_name} {member_user.last_name}" if member_user else "Unknown"
            
            print(f"\n💰 Processing debts for {member_name} (ID: {member_id})")
            
            # Get existing unpaid debts between member and payer
            existing_debts = db_manager.get_debts_between_users(member_id, payer_id, expense.board_id)
            unpaid_debts = [debt for debt in existing_debts if not debt.is_paid]
            
            # Calculate current debt balance
            member_owes_payer = sum(debt.amount for debt in unpaid_debts 
                                  if debt.from_user_id == member_id and debt.to_user_id == payer_id)
            payer_owes_member = sum(debt.amount for debt in unpaid_debts 
                                  if debt.from_user_id == payer_id and debt.to_user_id == member_id)
            
            print(f"💰 Current balance: {member_name} owes {member_owes_payer}, payer owes {payer_owes_member}")
            
            # Calculate net position: positive means member owes payer, negative means payer owes member
            current_balance = member_owes_payer - payer_owes_member
            # New debt: member owes payer amount_per_person (positive value)
            new_debt_to_payer = amount_per_person
            
            print(f"💰 Current net balance: {current_balance} (+ means {member_name} owes payer)")
            print(f"💰 New debt to payer from expense: {new_debt_to_payer}")
            
            # Calculate final balance: current + new debt to payer
            final_balance = current_balance + new_debt_to_payer
            print(f"💰 Final balance after expense: {final_balance}")
            
            # Now handle the offsetting logic - only offset when there are opposite direction debts
            if current_balance < 0 and new_debt_to_payer > 0:
                # Payer owes member money, new expense creates debt other direction → offset!
                amount_to_offset = min(abs(current_balance), new_debt_to_payer)
                print(f"💰 Can offset {amount_to_offset} - payer owes member, member owes from expense")
                
                # Offset existing debts where payer owes member
                offset_remaining = amount_to_offset
                payer_debts_to_member = [debt for debt in unpaid_debts 
                                       if debt.from_user_id == payer_id and debt.to_user_id == member_id]
                
                for debt in sorted(payer_debts_to_member, key=lambda x: x.amount):
                    if offset_remaining <= 0:
                        break
                    
                    # Get the actual DB debt object
                    db_debt = Debt.query.get(debt.id)
                    if not db_debt:
                        continue
                        
                    if debt.amount <= offset_remaining:
                        # Close this debt completely
                        print(f"💰 Closing payer debt {debt.id} completely (amount: {debt.amount}) - OFFSET!")
                        db_debt.is_paid = True
                        db_debt.paid_at = datetime.utcnow()
                        db_debt.paid_amount = db_debt.original_amount or db_debt.amount
                        if db_debt.original_amount is None:
                            db_debt.original_amount = db_debt.amount
                        offset_remaining -= debt.amount
                    else:
                        # Partially reduce this debt
                        old_amount = debt.amount
                        db_debt.amount -= offset_remaining
                        db_debt.paid_amount = (db_debt.paid_amount or 0) + offset_remaining
                        if db_debt.original_amount is None:
                            db_debt.original_amount = old_amount
                        print(f"💰 Partially reducing payer debt {debt.id}: {old_amount} → {db_debt.amount} - OFFSET!")
                        offset_remaining = 0
                
                # Create new debt only for the remaining amount after offset
                remaining_debt = new_debt_to_payer - amount_to_offset
                if remaining_debt > 0:
                    debt_data = {
                        'board_id': expense.board_id,
                        'expense_id': expense.id,
                        'from_user_id': member_id,
                        'to_user_id': payer_id,
                        'amount': remaining_debt,
                        'description': f"{expense.category} - {expense.description or 'Shared expense'}"
                    }
                    new_debt = db_manager.create_debt(debt_data)
                    print(f"💰 Created new debt {new_debt.id}: {member_name} owes {remaining_debt} to payer (after offset)")
                    debts_processed.append(new_debt.id)
                else:
                    print(f"💰 Perfect offset! No new debt needed - fully offset by existing debts")
            
            else:
                # No offsetting needed - just create new debt
                # This covers cases where:
                # 1. current_balance == 0 (no existing debts)
                # 2. current_balance > 0 (member already owes payer - debts accumulate)
                print(f"💰 No offsetting needed - creating new debt")
                debt_data = {
                    'board_id': expense.board_id,
                    'expense_id': expense.id,
                    'from_user_id': member_id,
                    'to_user_id': payer_id,
                    'amount': new_debt_to_payer,
                    'description': f"{expense.category} - {expense.description or 'Shared expense'}"
                }
                new_debt = db_manager.create_debt(debt_data)
                if current_balance > 0:
                    print(f"💰 Created new debt {new_debt.id}: {member_name} owes {new_debt_to_payer} to payer (accumulating with existing {current_balance})")
                else:
                    print(f"💰 Created new debt {new_debt.id}: {member_name} owes {new_debt_to_payer} to payer (no existing debts)")
                debts_processed.append(new_debt.id)

        
        # Commit all changes
        try:
            db_manager.db.session.commit()
            print(f"💰 Successfully processed debts for {len(other_members)} members")
            print(f"💰 Total new debts created: {len(debts_processed)}")
        except Exception as e:
            db_manager.db.session.rollback()
            print(f"💰 Error processing debts: {e}")
            raise
        
        print(f"💰 === Finished processing debts ===")
    
    # Keep the old function name for compatibility
    def create_debts_for_expense(expense, members):
        """Legacy function name - calls the new process_expense_debts"""
        return process_expense_debts(expense, members)

    # Notification endpoints
    @app.route('/api/notifications', methods=['GET'])
    @require_auth
    def get_notifications():
        """Get notifications for the current user"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        notifications = db_manager.get_user_notifications(current_user_id)
        
        notification_list = []
        for notification in notifications:
            notification_list.append({
                'id': notification.id,
                'user_id': notification.user_id,
                'board_id': notification.board_id,
                'board_name': notification.board_name,
                'expense_id': notification.expense_id,
                'expense_description': notification.expense_description,
                'amount': notification.amount,
                'created_by': notification.created_by,
                'created_by_name': notification.created_by_name,
                'created_at': notification.created_at,
                'is_read': notification.is_read,
                'type': notification.type
            })
        
        return jsonify({'notifications': notification_list}), 200

    @app.route('/api/notifications/<notification_id>/read', methods=['PUT'])
    @require_auth
    def mark_notification_as_read(notification_id):
        """Mark a specific notification as read"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        # Check if notification belongs to current user
        notifications = db_manager.get_user_notifications(current_user_id)
        notification = next((n for n in notifications if n.id == notification_id), None)
        
        if not notification:
            return jsonify({'error': 'Notification not found'}), 404
        
        success = db_manager.mark_notification_as_read(notification_id)
        
        if success:
            return jsonify({'message': 'Notification marked as read'}), 200
        else:
            return jsonify({'error': 'Failed to mark notification as read'}), 500

    @app.route('/api/notifications/mark-all-read', methods=['PUT'])
    @require_auth
    def mark_all_notifications_as_read():
        """Mark all notifications as read for the current user"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        success = db_manager.mark_all_user_notifications_as_read(current_user_id)
        
        if success:
            return jsonify({'message': 'All notifications marked as read'}), 200
        else:
            return jsonify({'error': 'Failed to mark notifications as read'}), 500

    @app.route('/api/notifications/<notification_id>', methods=['DELETE'])
    @require_auth
    def delete_notification(notification_id):
        """Delete a specific notification"""
        current_user_id = auth_manager.get_current_user()['user']['id']
        
        # Check if notification belongs to current user
        notifications = db_manager.get_user_notifications(current_user_id)
        notification = next((n for n in notifications if n.id == notification_id), None)
        
        if not notification:
            return jsonify({'error': 'Notification not found'}), 404
        
        success = db_manager.delete_notification(notification_id)
        
        if success:
            return jsonify({'message': 'Notification deleted'}), 200
        else:
            return jsonify({'error': 'Failed to delete notification'}), 500

    @app.route('/api/auth/check-terms-acceptance', methods=['GET'])
    @require_auth
    def check_terms_acceptance():
        """Check if current user has accepted the latest terms"""
        try:
            current_user = auth_manager.get_current_user()['user']
            user = db_manager.get_user_by_id(current_user['id'])
            
            if not user:
                return jsonify({'error': 'User not found'}), 404
            
            # Get latest terms version
            latest_terms = get_latest_terms_version()
            if not latest_terms:
                return jsonify({'error': 'No active terms version found'}), 500
            
            # Check if user has signed the latest version
            user_signed_version = user.terms_version_signed
            latest_version = latest_terms.version_number
            is_up_to_date = user_signed_version == latest_version
            
            return jsonify({
                'accepted_terms': user.accepted_terms,
                'terms_accepted_at': user.terms_accepted_at if user.terms_accepted_at else None,
                'terms_version_signed': user_signed_version,
                'current_terms_version': latest_version,
                'is_up_to_date': is_up_to_date,
                'requires_acceptance': not is_up_to_date
            }), 200
            
        except Exception as e:
            print(f"Error checking terms acceptance: {e}")
            return jsonify({'error': 'Failed to check terms acceptance'}), 500





    @app.route('/api/auth/accept-terms', methods=['POST'])
    @require_auth
    def accept_terms():
        """Mark current terms as accepted by user"""
        try:
            current_user = auth_manager.get_current_user()['user']
            user = db_manager.get_user_by_id(current_user['id'])
            
            if not user:
                return jsonify({'error': 'User not found'}), 404
            
            # Get latest terms version
            latest_terms = get_latest_terms_version()
            if not latest_terms:
                return jsonify({'error': 'No active terms version found'}), 500
            
            # Update terms acceptance
            update_data = {
                'accepted_terms': True,
                'terms_accepted_at': datetime.utcnow(),
                'terms_version_signed': latest_terms.version_number
            }
            
            success = db_manager.update_user(current_user['id'], update_data)
            
            if success:
                return jsonify({
                    'message': 'Terms accepted successfully',
                    'accepted_at': datetime.utcnow().isoformat(),
                    'terms_version_signed': latest_terms.version_number
                }), 200
            else:
                return jsonify({'error': 'Failed to update terms acceptance'}), 500
                
        except Exception as e:
            print(f"❌ Terms acceptance update error: {e}")
            return jsonify({'error': 'Failed to update terms acceptance'}), 500

    @app.route('/api/auth/terms-status', methods=['GET'])
    @require_auth
    def get_terms_status():
        """Get current user's terms acceptance status and latest version info"""
        try:
            current_user = auth_manager.get_current_user()['user']
            user = db_manager.get_user_by_id(current_user['id'])
            
            if not user:
                return jsonify({'error': 'User not found'}), 404
            
            # Get latest terms version
            latest_terms = get_latest_terms_version()
            
            if not latest_terms:
                return jsonify({'error': 'No active terms version found'}), 500
            
            # Check if user has signed the latest version
            user_signed_version = user.terms_version_signed
            latest_version = latest_terms.version_number
            is_up_to_date = user_signed_version == latest_version
            
            return jsonify({
                'user_terms_status': {
                    'accepted_terms': user.accepted_terms,
                    'terms_accepted_at': user.terms_accepted_at if user.terms_accepted_at else None,
                    'terms_version_signed': user_signed_version,
                    'is_up_to_date': is_up_to_date
                },
                'latest_terms': {
                    'version_number': latest_version,
                    'title': latest_terms.title,
                    'created_at': latest_terms.created_at.isoformat(),
                    'change_description': latest_terms.change_description
                },
                'requires_acceptance': not is_up_to_date
            }), 200
            
        except Exception as e:
            print(f"❌ Error getting terms status: {e}")
            return jsonify({'error': 'Failed to get terms status'}), 500

    @app.route('/api/terms/latest', methods=['GET'])
    def get_latest_terms_info():
        """Get information about the latest terms version (public endpoint)"""
        try:
            latest_terms = get_latest_terms_version()
            
            if not latest_terms:
                return jsonify({'error': 'No active terms version found'}), 404
            
            return jsonify({
                'version_number': latest_terms.version_number,
                'title': latest_terms.title,
                'created_at': latest_terms.created_at.isoformat(),
                'change_description': latest_terms.change_description,
                'endpoints': {
                    'hebrew': '/terms/he',
                    'english': '/terms/en',
                    'default': '/terms'
                }
            }), 200
            
        except Exception as e:
            print(f"❌ Error getting latest terms info: {e}")
            return jsonify({'error': 'Failed to get terms information'}), 500

    # Terms and Conditions API endpoints
    @app.route('/terms', methods=['GET'])
    def get_terms():
        """Get terms and conditions in Hebrew (default)"""
        return get_terms_html('he')
    
    @app.route('/terms/en', methods=['GET'])
    def get_terms_english():
        """Get terms and conditions in English"""
        return get_terms_html('en')
    
    @app.route('/terms/he', methods=['GET'])
    def get_terms_hebrew():
        """Get terms and conditions in Hebrew"""
        return get_terms_html('he')
    
    @app.route('/app-ads.txt', methods=['GET'])
    def get_app_ads_txt():
        """Return app-ads.txt content for Google AdMob verification"""
        app_ads_content = "google.com, pub-4216673023528064, DIRECT, f08c47fec0942fa0"
        return Response(app_ads_content, mimetype='text/plain')
    
    @app.route('/', methods=['GET'])
    def get_homepage():
        """Return homepage with information about HOMIS application"""
        html_content = """
        <!DOCTYPE html>
        <html lang="he" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>HOMIS - ניהול הוצאות משותפות</title>
            <style>
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    margin: 0;
                    padding: 0;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                }
                .container {
                    max-width: 1200px;
                    margin: 0 auto;
                    padding: 20px;
                }
                .header {
                    text-align: center;
                    color: white;
                    padding: 60px 0 40px 0;
                }
                .header h1 {
                    font-size: 3.5rem;
                    margin-bottom: 20px;
                    text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
                }
                .header p {
                    font-size: 1.3rem;
                    margin-bottom: 30px;
                    opacity: 0.9;
                }
                .content {
                    background: white;
                    border-radius: 15px;
                    padding: 40px;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                    margin-bottom: 30px;
                }
                .features {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
                    gap: 30px;
                    margin: 40px 0;
                }
                .feature {
                    background: #f8f9fa;
                    padding: 25px;
                    border-radius: 10px;
                    border-left: 4px solid #667eea;
                }
                .feature h3 {
                    color: #667eea;
                    margin-bottom: 15px;
                }
                .cta {
                    text-align: center;
                    background: #667eea;
                    color: white;
                    padding: 30px;
                    border-radius: 10px;
                    margin-top: 30px;
                }
                .cta h2 {
                    margin-bottom: 20px;
                }
                .cta p {
                    font-size: 1.1rem;
                    margin-bottom: 25px;
                }
                .download-btn {
                    display: inline-block;
                    background: #28a745;
                    color: white;
                    padding: 15px 30px;
                    text-decoration: none;
                    border-radius: 25px;
                    font-weight: bold;
                    transition: all 0.3s ease;
                }
                .download-btn:hover {
                    background: #218838;
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(0,0,0,0.2);
                }
                .footer {
                    text-align: center;
                    color: white;
                    padding: 20px 0;
                    opacity: 0.8;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🏠 HOMIS</h1>
                    <p>הפתרון המושלם לניהול הוצאות משותפות</p>
                </div>
                
                <div class="content">
                    <h2>מה זה HOMIS?</h2>
                    <p>HOMIS היא אפליקציה חכמה לניהול הוצאות משותפות בין חברים, משפחה או שותפים לדירה. 
                    האפליקציה מאפשרת לכם לעקוב אחר הוצאות משותפות, לחשב חובות ולשמור על שקיפות מלאה.</p>
                    
                    <div class="features">
                        <div class="feature">
                            <h3>📊 ניהול הוצאות</h3>
                            <p>הוסיפו הוצאות בקלות, קטגרו אותן ועקבו אחר ההוצאות המשותפות שלכם.</p>
                        </div>
                        <div class="feature">
                            <h3>💰 חישוב חובות</h3>
                            <p>המערכת מחשבת אוטומטית מי חייב למי ומציגה את החובות בצורה ברורה.</p>
                        </div>
                        <div class="feature">
                            <h3>👥 ניהול קבוצות</h3>
                            <p>צרו קבוצות עם חברים או שותפים וניהולו הוצאות משותפות בקלות.</p>
                        </div>
                        <div class="feature">
                            <h3>📱 גישה נוחה</h3>
                            <p>גישה לאפליקציה מכל מכשיר - טלפון, טאבלט או מחשב.</p>
                        </div>
                        <div class="feature">
                            <h3>🔒 אבטחה מתקדמת</h3>
                            <p>המידע שלכם מוגן ומאובטח עם מערכת אבטחה מתקדמת.</p>
                        </div>
                        <div class="feature">
                            <h3>📈 דוחות מפורטים</h3>
                            <p>קבלו דוחות מפורטים על ההוצאות והחובות שלכם.</p>
                        </div>
                    </div>
                    
                    <div class="cta">
                        <h2>מוכנים להתחיל?</h2>
                        <p>הורידו את האפליקציה עכשיו והתחילו לנהל את ההוצאות המשותפות שלכם בצורה חכמה!</p>
                        <a href="#" class="download-btn">📱 הורדת האפליקציה</a>
                    </div>
                </div>
                
                <div class="footer">
                    <p>&copy; 2024 HOMIS - כל הזכויות שמורות</p>
                </div>
            </div>
        </body>
        </html>
        """
        return Response(html_content, mimetype='text/html')
    
    def get_latest_terms_version():
        """Get the latest active terms version from database"""
        try:
            latest_terms = TermsVersion.query.filter_by(is_active=True).order_by(TermsVersion.version_number.desc()).first()
            return latest_terms
        except Exception as e:
            print(f"Error getting latest terms: {e}")
            return None

    def get_terms_html(language='he'):
        """Get HTML terms and conditions from database based on language"""
        # Get latest terms from database
        latest_terms = get_latest_terms_version()
        
        if not latest_terms:
            # Fallback to default message if no terms in database
            return Response("""
            <!DOCTYPE html>
            <html>
            <head><title>Terms Not Available</title></head>
            <body>
                <h1>Terms and Conditions</h1>
                <p>Terms and conditions are not available at this time. Please contact support.</p>
            </body>
            </html>
            """, mimetype='text/html')
        
        # Get content based on language
        if language == 'en':
            html_content = latest_terms.content_english
        else:
            html_content = latest_terms.content_hebrew
        
        return Response(html_content, mimetype='text/html')

    def get_legacy_terms_html(language='he'):
        """Legacy function - Generate static HTML terms and conditions based on language (for reference)"""
        if language == 'en':
            html_content = f"""
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>Terms of Service - Homis</title>
                <style>
                    body {{
                        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                        line-height: 1.6;
                        color: #333;
                        max-width: 800px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .container {{
                        background: white;
                        padding: 30px;
                        border-radius: 10px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    }}
                    h1 {{
                        color: #2c3e50;
                        text-align: center;
                        border-bottom: 3px solid #3498db;
                        padding-bottom: 10px;
                    }}
                    h2 {{
                        color: #34495e;
                        margin-top: 30px;
                        margin-bottom: 15px;
                        border-left: 4px solid #3498db;
                        padding-left: 15px;
                    }}
                    p, ul {{
                        margin-bottom: 15px;
                        text-align: justify;
                    }}
                    ul {{
                        padding-left: 20px;
                    }}
                    li {{
                        margin-bottom: 8px;
                    }}
                    .last-updated {{
                        text-align: center;
                        color: #7f8c8d;
                        font-style: italic;
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 1px solid #ecf0f1;
                    }}
                    .language-toggle {{
                        text-align: center;
                        margin-bottom: 20px;
                    }}
                    .language-toggle a {{
                        display: inline-block;
                        padding: 10px 20px;
                        background: #3498db;
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin: 0 10px;
                        transition: background 0.3s;
                    }}
                    .language-toggle a:hover {{
                        background: #2980b9;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="language-toggle">
                        <a href="/api/terms/he">עברית</a>
                        <a href="/api/terms/en" style="background: #27ae60;">English</a>
                    </div>
                    
                    <h1>Terms of Service</h1>
                    
                    <h2>1. Acceptance of Terms</h2>
                    <p>Using the "Shared Expenses Management" application constitutes agreement to the terms of use detailed below. If you do not agree to these terms, please do not use the application.</p>
                    
                    <h2>2. Service Description</h2>
                    <p>The application allows users to manage shared household expenses, track spending, manage debts, and share financial information with family members or roommates.</p>
                    
                    <h2>3. Registration and User Account</h2>
                    <ul>
                        <li>The user must provide accurate and up-to-date information during registration</li>
                        <li>The user is responsible for maintaining the confidentiality of their login credentials</li>
                        <li>The user is responsible for all activity conducted under their account</li>
                        <li>Sharing a user account with others is prohibited</li>
                    </ul>
                    
                    <h2>4. Privacy and Security</h2>
                    <ul>
                        <li>We are committed to protecting user privacy in accordance with our privacy policy</li>
                        <li>Financial information is stored on secure servers</li>
                        <li>However, we cannot guarantee absolute security against breaches or technical failures</li>
                    </ul>
                    
                    <h2>5. User Responsibility</h2>
                    <ul>
                        <li>The user is responsible for the accuracy of information entered into the system</li>
                        <li>The user agrees not to use the application for illegal purposes</li>
                        <li>The user will not misuse the system or harm other users</li>
                    </ul>
                    
                    <h2>6. Limitation of Liability</h2>
                    <ul>
                        <li>The application is provided "as is" without any warranty</li>
                        <li>We are not responsible for any direct or indirect damage that may result from using the application</li>
                        <li>We do not guarantee that the service will be available without interruption or without errors</li>
                    </ul>
                    
                    <h2>7. Service Termination</h2>
                    <ul>
                        <li>The application developer may decide to stop developing the application and shut down the servers at any given moment</li>
                        <li>Users agree that they cannot claim compensation for not having access to information they stored</li>
                        <li>Information storage is for organizational purposes only, and we do not commit to complete and long-term storage</li>
                        <li>Users are responsible for independently backing up their important information</li>
                    </ul>
                    
                    <h2>8. Changes to Terms</h2>
                    <p>We may update the terms of use from time to time. Material changes will be communicated to users. Continued use of the application after changes constitutes agreement to the new terms.</p>
                    
                    <h2>9. Intellectual Property</h2>
                    <p>All rights to the application, including code, design, and content, belong to the application developer. Copying, duplicating, or distributing the application without explicit permission is prohibited.</p>
                    
                    <h2>10. Governing Law</h2>
                    <p>These terms of use are subject to the laws of the State of Israel. Any dispute will be resolved before the competent court in Israel.</p>
                    
                    <h2>11. Contact</h2>
                    <p>For questions or complaints regarding the terms of use, you can contact the application developer.</p>
                    
                    <div class="last-updated">
                        Last updated: {datetime.now().strftime('%B %d, %Y')}
                    </div>
                </div>
            </body>
            </html>
            """
        else:
            # Hebrew version
            html_content = f"""
            <!DOCTYPE html>
            <html dir="rtl" lang="he">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>תנאי שימוש - Homis</title>
                <style>
                    body {{
                        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                        line-height: 1.6;
                        color: #333;
                        max-width: 800px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .container {{
                        background: white;
                        padding: 30px;
                        border-radius: 10px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    }}
                    h1 {{
                        color: #2c3e50;
                        text-align: center;
                        border-bottom: 3px solid #3498db;
                        padding-bottom: 10px;
                    }}
                    h2 {{
                        color: #34495e;
                        margin-top: 30px;
                        margin-bottom: 15px;
                        border-right: 4px solid #3498db;
                        padding-right: 15px;
                    }}
                    p, ul {{
                        margin-bottom: 15px;
                        text-align: justify;
                    }}
                    ul {{
                        padding-right: 20px;
                    }}
                    li {{
                        margin-bottom: 8px;
                    }}
                    .last-updated {{
                        text-align: center;
                        color: #7f8c8d;
                        font-style: italic;
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 1px solid #ecf0f1;
                    }}
                    .language-toggle {{
                        text-align: center;
                        margin-bottom: 20px;
                    }}
                    .language-toggle a {{
                        display: inline-block;
                        padding: 10px 20px;
                        background: #3498db;
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin: 0 10px;
                        transition: background 0.3s;
                    }}
                    .language-toggle a:hover {{
                        background: #2980b9;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="language-toggle">
                        <a href="/terms/he" style="background: #27ae60;">עברית</a>
                        <a href="/terms/en">English</a>
                    </div>
                    
                    <h1>תנאי שימוש</h1>
                    
                    <h2>1. קבלת התנאים</h2>
                    <p>השימוש באפליקציה "ניהול הוצאות משותפות" מהווה הסכמה לתנאי השימוש המפורטים להלן. אם אינך מסכים לתנאים אלה, אנא אל תשתמש באפליקציה.</p>
                    
                    <h2>2. תיאור השירות</h2>
                    <p>האפליקציה מאפשרת למשתמשים לנהל הוצאות משותפות בבית, לעקוב אחר הוצאות, לנהל חובות, ולשתף מידע פיננסי עם חברי המשפחה או השותפים.</p>
                    
                    <h2>3. הרשמה וחשבון משתמש</h2>
                    <ul>
                        <li>על המשתמש לספק מידע מדויק ומעודכן בעת ההרשמה</li>
                        <li>המשתמש אחראי לשמירת סודיות פרטי ההתחברות שלו</li>
                        <li>המשתמש אחראי לכל הפעילות המתבצעת בחשבונו</li>
                        <li>אסור לשתף חשבון משתמש עם אחרים</li>
                    </ul>
                    
                    <h2>4. פרטיות ואבטחה</h2>
                    <ul>
                        <li>אנו מתחייבים להגן על פרטיות המשתמשים בהתאם למדיניות הפרטיות שלנו</li>
                        <li>המידע הפיננסי נשמר בשרתים מאובטחים</li>
                        <li>עם זאת, אין אנו יכולים להבטיח אבטחה מוחלטת מפני פריצות או תקלות טכניות</li>
                    </ul>
                    
                    <h2>5. אחריות המשתמש</h2>
                    <ul>
                        <li>המשתמש אחראי לדיוק המידע שהוא מזין למערכת</li>
                        <li>המשתמש מסכים לא להשתמש באפליקציה למטרות בלתי חוקיות</li>
                        <li>המשתמש לא יעשה שימוש לרעה במערכת או יפגע במשתמשים אחרים</li>
                    </ul>
                    
                    <h2>6. הגבלת אחריות</h2>
                    <ul>
                        <li>האפליקציה מסופקת "כפי שהיא" ללא כל אחריות</li>
                        <li>אנו לא אחראים לכל נזק ישיר או עקיף שיכול להיגרם מהשימוש באפליקציה</li>
                        <li>אנו לא מתחייבים שהשירות יהיה זמין ללא הפרעה או ללא שגיאות</li>
                    </ul>
                    
                    <h2>7. הפסקת השירות</h2>
                    <ul>
                        <li>מפתח האפליקציה רשאי להחליט על הפסקת פיתוח האפליקציה וסגירת השרתים בכל רגע נתון</li>
                        <li>המשתמשים מסכימים שלא יוכלו לתבוע על כך שאין להם גישה למידע ששמרו</li>
                        <li>שמירת המידע היא לצורך סדר וארגון בלבד, ואנו איננו מתחייבים לשמירה מלאה ולאורך זמן</li>
                        <li>המשתמשים אחראים לגיבוי המידע החשוב שלהם באופן עצמאי</li>
                    </ul>
                    
                    <h2>8. שינויים בתנאים</h2>
                    <p>אנו רשאים לעדכן את תנאי השימוש מעת לעת. שינויים מהותיים יובאו לידיעת המשתמשים. המשך השימוש באפליקציה לאחר השינויים מהווה הסכמה לתנאים החדשים.</p>
                    
                    <h2>9. קניין רוחני</h2>
                    <p>כל הזכויות באפליקציה, כולל הקוד, העיצוב והתוכן, שייכות למפתח האפליקציה. אסור להעתיק, לשכפל או להפיץ את האפליקציה ללא אישור מפורש.</p>
                    
                    <h2>10. דין שולט</h2>
                    <p>תנאי שימוש אלה כפופים לחוקי מדינת ישראל. כל מחלוקת תיפתר בפני בית המשפט המוסמך בישראל.</p>
                    
                    <h2>11. יצירת קשר</h2>
                    <p>לשאלות או תלונות בנוגע לתנאי השימוש, ניתן ליצור קשר עם מפתח האפליקציה.</p>
                    
                    <div class="last-updated">
                        עודכן לאחרונה: {datetime.now().strftime('%d/%m/%Y')}
                    </div>
                </div>
            </body>
            </html>
            """
        
        return Response(html_content, mimetype='text/html')

    @app.route('/api/auth/delete-user', methods=['DELETE'])
    @require_auth
    def delete_user():
        """Delete current user account and all related data"""
        try:
            current_user = auth_manager.get_current_user()['user']
            user_id = current_user['id']
            
            print(f"🗑️ User {user_id} ({current_user['email']}) requested account deletion")
            
            # Get user info for logging
            user = db_manager.get_user_by_id(user_id)
            if not user:
                print(f"❌ User {user_id} not found in database")
                return jsonify({'error': 'User not found'}), 404
            
            print(f"🗑️ Found user in database: {user.email}")
            
            # Get user's boards and data for logging
            user_boards = db_manager.get_user_boards(user_id)
            owned_boards = [board for board in user_boards if board.owner_id == user_id]
            
            print(f"🗑️ User has {len(user_boards)} board memberships, owns {len(owned_boards)} boards")
            
            # Delete user and all related data
            print(f"🗑️ Calling db_manager.delete_user({user_id})")
            success = db_manager.delete_user(user_id)
            print(f"🗑️ db_manager.delete_user returned: {success}")
            
            if success:
                # Verify the user was actually deleted
                print(f"🗑️ Verifying user deletion...")
                deleted_user = db_manager.get_user_by_id(user_id)
                if deleted_user:
                    print(f"❌ User still exists after deletion! ID: {deleted_user.id}, Email: {deleted_user.email}")
                    return jsonify({'error': 'User deletion failed - user still exists'}), 500
                else:
                    print(f"✅ User {user_id} ({current_user['email']}) verified as deleted")
                    return jsonify({'message': 'User account deleted successfully'}), 200
            else:
                print(f"❌ Failed to delete user {user_id}")
                return jsonify({'error': 'Failed to delete user account'}), 500
                
        except Exception as e:
            print(f"❌ Error deleting user: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': 'Failed to delete user account'}), 500

    return app

if __name__ == '__main__':
    app = create_app(os.getenv('FLASK_ENV', 'development'))
    app.run(debug=app.config['DEBUG'], host='0.0.0.0', port=5000,use_reloader=False)

